import os
import sys
import urllib
import urllib2
from bs4 import BeautifulSoup, Tag, NavigableString, Comment
import collections
import math
import csv
import codecs
import json
import inspect
import urlparse
import argparse
import traceback
import time
import datetime
import openpyxl

import Main_Extractor as main_ext

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.firefox.options import Options

from common import shared
from common import bsoup
from common import access_rest as rest
from common import page_group
from common import recurse_ftp as rec_ftp
from common import spreadsheet as sh

class PT_Extractor(main_ext.Extractor):
	def __init__(self):
		''' Initializer for the Extractor class. '''

		# Set the province
		self.province = 'Ontario'
		
		# Create the page groups dictionary
		self.page_groups = []
		self.page_groups = collections.OrderedDict()

		# Declare all the different types of pages
		pg_grp = page_group.PageGroup('opendata', "Ontario Open Data Catalogue")
		pg_grp.add_url('main_url', 'https://www.ontario.ca/search/data-catalogue')
		opts = {'format': ['csv', 'json', 'kml', 'mdb', 'txt', 'xls', 'xlsx', 'zip'],
				'status': ['open', 'to be opened', 'under review', 'restricted']}
		pg_grp.set_opts(opts)
		pg_grp.add_default('status', 'open')
		self.page_groups['opendata'] = pg_grp

		pg_grp = page_group.PageGroup('discovering', "Discovering Ontario")
		pg_grp.add_url('main_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en/main.home')
		pg_grp.add_url('portal_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en/main.search.embedded')
		pg_grp.add_url('mdata_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en/metadata.show.embedded')
		self.page_groups['discovering'] = pg_grp
		
		pg_grp = page_group.PageGroup('update', "Update Discovering Ontario")
		pg_grp.add_url('mdata_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en/metadata.show.embedded')
		self.page_groups['update'] = pg_grp

		# Initialize the Main Extractor to use its variables
		main_ext.Extractor.__init__(self)
		
		# Set the arguments for this extractor
		self.argmt['word'] = main_ext.Ext_Arg('word', 
										methlst=['opendata', 'discovering'])
		self.argmt['format'] = main_ext.Ext_Arg('format', methlst=['opendata'])
		self.argmt['status'] = main_ext.Ext_Arg('status', methlst=['opendata'])
		self.argmt['downloadable'] = main_ext.Ext_Arg('downloadable', 
												methlst=['discovering'])
		sp_opts = ['imagery', 'wms', 'mp3', 'application', 'maps', 'liowarehouse', 
					'publicwebservice', 'gisdata', 'opendata', 'non-commercialuse', 
					'geologyontario', 'elevation']
		self.argmt['subpage'] = main_ext.Ext_Arg('subpage', 
												methlst=['discovering'], 
												opts=sp_opts)
		
	def get_province(self):
		''' Gets the province name of the extractor.
		:return: The province name of the extractor.
		'''
		
		return self.province

	###################################################################################################################

	def get_discovering_mdata(self, mdata_soup, url, title_str):
	
		if self.check_result(mdata_soup, url, "Metadata Page - %s" % title_str, output=False):
				
			# Get date from adjacent element of <th> with text 'Date'
			date_str = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Date', url=url)

			# Get description from adjacent element of <th> with text 'Abstract'
			desc_str = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Abstract', url=url)
			#print desc_str

			# Get data type from adjacent element of <th> with text 'Environment description'
			dtype_str = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Environment description', url=url)

			# Get the publisher from adjacent element of <th> with text 'Organisation name'
			pub_str = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Organisation name', url=url)

			# Get Data URL(s)
			download_items = mdata_soup.find_all('span', string='Data for download')
			data_links = []
			for i in download_items:
				th_html = i.parent
				td_html = th_html.find_next_sibling('td')
				data_link = td_html.a['href']
				data_links.append(data_link)

			# Get metadata type from adjacent element of <th> with text 'Metadata standard name'
			mdata_type = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Metadata standard name', url=url)

			# Get licence from adjacent element of <th> with text 'Access constraints'
			#lic_str = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Access constraints')
			tags = [('th', 'Use limitation'), ('th', 'Other constraints'), ('th', 'Access constraints')]
			lic_str = self.get_mdata(mdata_soup, tags)
			#print mdata_url
			#print "lic_str: %s" % lic_str

			# Get the spatial reference
			ref_span = mdata_soup.find('span', string='Reference System Information')
			sp_str = ''
			if ref_span is not None:
				ref_td = bsoup.get_parent(ref_span, 'td')
				sp_str = bsoup.get_adj_text_by_label(ref_td, 'th', 'Code', url=url)

			# Fill in the access and download_str variables based on the number of
			#   data_links
			if len(data_links) == 0:
				download_str = 'No'
				access = 'Contact the Province'
			elif len(data_links) == 1:
				download_str = data_links[0]
				access = 'Download/Web Accessible'
			else:
				download_str = 'Multiple Downloads'
				access = 'Download/Web Accessible'
		else:
			desc_str = ''
			pub_str = ''
			dtype_str = ''
			lic_str = ''
			date_str = ''
			access = ''
			download_str = ''
			sp_str = ''
			mdata_type = ''
			#mdata_url = ''
			
		# FOR DEBUG:
		# if self.debug:
			# print "\n"
			# print "Title: %s" % title_str
			# print "Description: %s" % desc_str
			# print "Publisher: %s" % pub_str
			# print "Type: %s" % dtype_str
			# print "Licensing: %s" % lic_str
			# print "Date: %s" % date_str
			# print "Access: %s" % access
			# print "Download: %s" % download_str
			# print "Spatial Reference: %s" % sp_str
			# print "Metadata Type: %s" % mdata_type
			# print "Metadata URL: %s" % url
			# print "Notes: %s" % self.notes
			# answer = raw_input("Press enter...")

		mdata_items = collections.OrderedDict()
			
		mdata_items['Title'] = title_str
		mdata_items['Description'] = desc_str
		mdata_items['Publisher'] = pub_str
		#pt_csv.add('Data URL'] = '|'.join(data_links)
		mdata_items['Type'] = dtype_str
		mdata_items['Licensing'] = lic_str
		mdata_items['Date'] = date_str
		mdata_items['Access'] = access
		mdata_items['Download'] = download_str
		mdata_items['Spatial Reference'] = sp_str
		mdata_items['Metadata Type'] = mdata_type
		mdata_items['Metadata URL'] = url
		mdata_items['Notes'] = self.notes
		
		return mdata_items
		
	def get_category(self, cat_div):
		# Get the onclick text to get the category
		cat_text = cat_div['onclick']
		#print "cat_text: %s" % cat_text
		start_pos = cat_text.find("(")
		end_pos = cat_text.find(")", start_pos)
		category = cat_text[start_pos+2:end_pos-1]
		#print "category: %s" % category
		
		return category
	
	def get_next_twenty(self, driver):
		''' Gets the element with text 'Next 20'
		:param driver: The Selenium driver
		:return: Element with text 'Next 20'
		'''
		found_elements = driver.find_elements_by_class_name("secondary")
		for el in found_elements:
			if el.text.strip() == 'Next 20':
				return el
				
	def get_mdata(self, mdata_soup, heading_tups):
		
		mdata_val = ''
		for heading in heading_tups:
			if mdata_val == '':
				tag = heading[0]
				name = heading[1]
				mdata_val = bsoup.get_adj_text_by_label(mdata_soup, tag, name)
				
		return mdata_val

	def extract_opendata(self): #, word=None, format=None, status=None):
		''' Extracts the results from the Open Data Catalogue of Ontario
		:param word: The word used to filter the results
		:param format: The format used to filter the results
		:param status: The status used to filter the results
		:return: None
		'''
		
		# Get the parameters
		word = self.argmt['word'].get_value()
		format = self.argmt['format'].get_value()
		status = self.argmt['status'].get_value()

		# URL examples for statuses:
		# https://www.ontario.ca/search/data-catalogue?sort=asc&status=%5B%2230%22%5D
		# https://www.ontario.ca/search/data-catalogue?sort=asc&status=["30"]

		# Full example of possible URL:
		# https://www.ontario.ca/search/data-catalogue?
		# sort=asc
		# &query=gis
		# &filetype=%5B%22csv%22%5D
		# &publisher=%5B%22Advanced%20Education%20and%20Skills%20Development%22%5D
		# &topic=%5B%22Arts%20and%20culture%22%5D
		# &status=%5B%2230%22%5D

		# Different statuses:
		# 30 - Open
		# 20 - To be opened
		# 10 - Under review
		# 0 - Restricted

		main_url = self.pg_grp.get_url('main_url')

		self.print_log("\nExtracting from %s" % self.pg_grp.get_title())
		
		self.print_title("Extracting Ontario's Open Data Catalogue")
		
		start_time = datetime.datetime.now()
		print "Process started at: %s" % start_time

		# Create the CSV file
		csv_fn = "OpenData_results"
		pt_csv = sh.PT_CSV(csv_fn, self)
		pt_csv.open_csv()

		# Set the parameters for the URL query
		params = collections.OrderedDict()
		if word is not None and not word == '':
			params['query'] = word
		if format is not None and not format == '':
			params['filetype'] = '["%s"]' % format
		if status is not None and not status == '':
			if status.lower() == 'open':
				status_val = '30'
			elif status.lower() == 'to be opened':
				status_val = '20'
			elif status.lower() == 'under review':
				status_val = '10'
			else:
				status_val = '0'
			params['status'] = '["%s"]' % status_val

		# Build the URL query
		query_url = shared.get_post_query(main_url, params)

		print "\nQuery URL: %s" % query_url

		# Create a Selenium driver with a headless Firefox
		#options = Options()
		#options.add_argument("--headless")
		driver = webdriver.Firefox()#firefox_options=options)
		driver.implicitly_wait(20)
		driver.get(query_url)

		# Wait for the element with ID "rechercheEdelForm:lancerRecherche" to load
		shared.wait_page_load(driver, By.CLASS_NAME, "results-page")

		# Continually click 'Next 20' button until it no longer exists
		next_20 = self.get_next_twenty(driver)
		while next_20 is not None:
			next_20.click()
			driver.implicitly_wait(20)
			next_20 = self.get_next_twenty(driver)

		# Grab the results of the results page once all results are shown
		page_html = driver.page_source

		# Create the BeautifulSoup object
		soup = BeautifulSoup(page_html, 'html.parser')

		# Get all the datasets on the current page (all datasets are in a 'div' with class 'dataset-item')
		page_list = soup.find_all('ul', attrs={'class': 'results-page'})

		# Get the number of records for status purposes
		res_span = bsoup.find_tags_containing(soup, 'results', 'span')
		res_text = res_span.text
		res_text = res_text.strip()
		end_pos = res_text.find(' results')
		num_results = res_text[:end_pos]

		print "The number of results is %s" % num_results

		record_count = 0

		# Cycle through each page DIV
		for page in page_list:

			li_res = page.find_all('li')

			for li in li_res:

				record_count += 1

				#print "Extracting record %s of %s" % (record_count, num_results)
				msg = "Extracting record %s of %s" % (record_count, num_results)
				shared.print_oneliner(msg)

				# Get the title and URL
				h3 = li.find('h3')
				a = h3.find('a')
				title_str = a.contents[0].strip()
				mdata_href = a['href']
				mdata_url = urlparse.urljoin(query_url, mdata_href)

				date_str = ''
				pub_str = ''
				dt_list = li.find_all('dt')
				for dt in dt_list:
					# Get the date
					if dt.text.find("Date added") > -1:
						dd_html = dt.find_next_sibling('dd')
						date_str = dd_html.string.strip()

					# Get the publisher
					elif dt.text.find("Publisher") > -1:
						dd_html = dt.find_next_sibling('dd')
						pub_str = dd_html.text.strip()

				# Get the rest of the values using the metadata URL
				attrb = ('class', 'thumbs-down')
				mdata_soup = bsoup.get_soup(mdata_url, True, attrb)
				
				if self.check_result(mdata_soup, mdata_url, 
					"Metadata Page - %s" % title_str, output=False):

					# Get the description
					div_pgbody = mdata_soup.find('div', attrs={'id': 'pagebody'})
					desc_str = bsoup.get_text(div_pgbody)

					# Get the downloads
					h2 = bsoup.get_adj_tags_by_text(mdata_soup, 'h2', 'Download data', True)
					if len(h2) == 0:
						download_str = 'No'
						access = 'Contact the Province'
						format_list = []
					else:
						dwnload_div = h2[0].parent
						anchors = dwnload_div.find_all('a')
						if len(anchors) == 0:
							download_str = 'No'
							access = 'Contact the Province'
						elif len(anchors) == 1:
							download_str = anchors[0]['href']
							access = "Download/Web Accessible"
						else:
							download_str = 'Multiple Downloads'
							access = "Download/Web Accessible"

						# Get the formats from the downloads
						format_list = []
						for a in anchors:
							format = a.abbr.text
							format_list.append(format)
							
				else:
					# Set the variables as empty strings if the metadata can't be loaded
					desc_str = ''
					date_str = ''
					pub_str = ''
					access = ''
					download_str = ''
					format_list = []
					mdata_url = ''

				# Add all values to the CSV file
				pt_csv.add('Title', title_str)
				pt_csv.add('Description', shared.edit_description(desc_str))
				pt_csv.add('Date', date_str)
				pt_csv.add('Publisher', pub_str)
				pt_csv.add('Access', access)
				pt_csv.add('Download', download_str)
				#pt_csv.add('Licensing', lic_str)
				pt_csv.add('Available Formats', '|'.join(format_list))
				pt_csv.add('Metadata URL', mdata_url)
				pt_csv.add('Notes', self.notes)

				# Write the results to the CSV
				pt_csv.write_dataset()
				
				self.notes = ''

		driver.quit()

		pt_csv.close_csv()
		
		# Print ending time
		end_time = datetime.datetime.now()
		print "\nExtraction complete at %s." % end_time
		tot_time = end_time - start_time
		print "It took %s to complete." % tot_time

	def extract_discovering(self): #, word=None, downloadable=False):
		''' Extracts results from the Land Information Ontario geoportal
		:param word: The word used to filter the results
		:param downloadable: Determines whether to include results only with downloads
		:return: None
		'''
		
		# Get the parameters
		word = self.argmt['word'].get_value()
		subpage = self.argmt['subpage'].get_value()
		downloadable = self.argmt['downloadable'].get_value()

		self.print_log("\nExtracting from %s" % self.pg_grp.get_title())
		
		self.print_title("Extracting Discovering Ontario Data")
		
		start_time = datetime.datetime.now()
		print "Process started at: %s" % start_time

		main_url = self.pg_grp.get_url('main_url')
		portal_url = self.pg_grp.get_url('portal_url')
		mdata_query_url = self.pg_grp.get_url('mdata_url')

		# Load the main page to get a list of categories
		attrb = ['id', 'latest_updates']
		main_soup = bsoup.get_soup(main_url, True, attrb)
		
		print "\nMain URL: %s" % main_url
		
		if not self.check_result(main_soup, main_url, "Ontario Discover Portal"): return None

		divs = main_soup.find_all('div')

		# Find all <div> within <div> with class 'geosearchfields' to get the different categories
		cat_divs = []
		for div in divs:
			if div.has_attr('class'):
				#print div['class']
				if div['class'][0].find('geosearchfields') > -1:
					cat_divs = div.find_all('div')
					break
					
		# Convert the category list of elements to words
		categories = [self.get_category(c) for c in cat_divs]
		
		print "subpage: %s" % subpage
		print "categories: %s" % categories
		# If the subpage is in the available categories,
		#	set the list of categories to the subpage
		for c in categories:
			if subpage.lower() == c.lower():
				categories = [c]
				break

		for category in categories:
			
			# Create the CSV file
			csv_fn = "%s_results" % category.title()
			pt_csv = sh.PT_CSV(csv_fn, self)
			#if pt_csv.check_exists():
			#	print "CSV file for category '%s' already exists." \
			#		  " Proceeding to next category." % category
			pt_csv.open_csv()

			# Set the parameters for the URL query
			params = collections.OrderedDict()
			if word is not None and not word == "": params['any'] = word
			params['category'] = category
			if downloadable: params['download'] = 'on'
			params['hitsPerPage'] = '1000'

			# Build the URL query
			query_url = shared.get_post_query(portal_url, params)

			print "\nQuery URL: %s" % query_url

			# Get the soup of the query URL
			xml_soup = bsoup.get_xml_soup(query_url)
			
			if not self.check_result(xml_soup, query_url, "Ontario Discover Portal Query"): continue

			items = xml_soup.find_all('div', attrs={'class': 'hit'})

			num_items = len(items)

			print "\nNumber of records: " + str(num_items)

			record_count = 0

			for index, item in enumerate(items):
			
				record_count += 1
				msg = "Extracting %s of %s records for category '%s'" % (record_count, num_items, category)
				shared.print_oneliner(msg)

				# Get the ID from <div> with class 'thumbnail_results'
				thumb_div = item.find('div', attrs={'class': 'thumbnail_results'})
				rating_link = thumb_div.a['id']
				id_str = rating_link.replace("rating.link.", "")
				#print "ID: %s" % id_str

				# Get the title from <div> with class 'hittitle'
				title_str = item.find('div', attrs={'class': 'hittitle'}).text
				
				#answer = raw_input("Press enter...")

				# Metadata URL:
				mdata_url = '%s?id=%s' % (mdata_query_url, id_str)

				# HTML Access
				mdata_soup = bsoup.get_soup(mdata_url) #, True, ('class', 'padded-content'))
				
				mdata_items = self.get_discovering_mdata(mdata_soup, mdata_url, title_str)
				
				for k, v in mdata_items.items():
					pt_csv.add(k, v)

				# Write the results to the CSV
				pt_csv.write_dataset()

			print

		pt_csv.close_csv()
		
		# Print ending time
		end_time = datetime.datetime.now()
		print "\nExtraction complete at %s." % end_time
		tot_time = end_time - start_time
		print "It took %s to complete." % tot_time
		
	def update_discover(self):
		''' Extracts the latest Discovering Ontario information
		
		'''
		
		self.print_title("Updating Discovering Ontario Data")
		
		start_time = datetime.datetime.now()
		print "Process started at: %s" % start_time
		
		excel_url = 'https://www.sse.gov.on.ca/sites/MNR-PublicDocs/EN/CMID/DataDistributionCatalogue.xlsx'
		
		mdata_query_url = self.pg_grp.get_url('mdata_url')

		now = datetime.datetime.now()
		date_str = now.strftime('%Y%m%d')

		# Download the XLSX file
		xlsx_fn = 'files\\DataDistributionCatalogue_Update_%s.xlsx' % date_str
		xlsx_f = urllib.urlretrieve(excel_url, xlsx_fn)

		#r = requests.get(excel_url)  # make an HTTP request

		wb = openpyxl.load_workbook(xlsx_fn)

		sheets = wb.sheetnames
		
		url_list = []

		for sh_idx, sh_name in enumerate(sheets):
			print "\nSheet %s:" % sh_name

			sheet = wb[sh_name]
			
			row_count = sheet.max_row
			
			if sh_idx == 0:
				row_start = 5
			else:
				row_start = 1
				
			for row in range(row_start, row_count - row_start + 1):
				cell = sheet['A%s' % row]
				hp_link = cell.hyperlink
				if hp_link is not None:
					#print hp_link
					title_str = cell.value
					url = hp_link.target
					if url is not None:
						print url
						
						# Parse the URL to get the ID
						id_str = url.split('=')[1]
						
						# Metadata URL:
						mdata_url = '%s?uuid=%s' % (mdata_query_url, id_str)
						
						print mdata_url
						
						url_list.append((title_str, mdata_url))
						
				#answer = raw_input("Press enter...")
						
		#answer = raw_input("Press enter...")
						
		# Create the CSV file
		csv_fn = "Discovering_Update_%s_results" % date_str
		pt_csv = sh.PT_CSV(csv_fn, self)
		pt_csv.open_csv()
		
		for idx, info in enumerate(url_list):
			msg = "Extracting %s of %s records" % (idx + 1, len(url_list))
			shared.print_oneliner(msg)
			
			title_str, url = info
		
			mdata_soup = bsoup.get_soup(url, True)
			
			#mdata_f = codecs.open('mdata_soup.html', encoding='utf-8', mode='w')
			#mdata_f.write(unicode(mdata_soup))
			#mdata_f.close()
			#mdata_f.close()
			
			mdata_items = self.get_discovering_mdata(mdata_soup, url, title_str)

			for k, v in mdata_items.items():
				pt_csv.add(k, v)

			# Write the results to the CSV
			pt_csv.write_dataset()

		print

		pt_csv.close_csv()
		
		# Print ending time
		end_time = datetime.datetime.now()
		print "\nExtraction complete at %s." % end_time
		tot_time = end_time - start_time
		print "It took %s to complete." % tot_time

def main():
	# parser.add_argument("-t", "--tool", help="The tool to use: %s" % ', '.join(tool_list))
	# parser.add_argument("-w", "--word", help="The key word(s) to search for.")
	# parser.add_argument("-f", "--format", help="The format(s) to search for.")
	# parser.add_argument("-c", "--category", help="The category(ies) to search for.")
	# parser.add_argument("-d", "--downloadable", help="Determines wheter to get only downloadable datasets.")
	# parser.add_argument("-l", "--html", help="The HTML file to scrape (only for OpenData website).")
	# parser.add_argument("-s", "--silent", action='store_true', help="If used, no extra parameters will be queried.")

	ext = Extractor()

	try:
		pages = ext.get_pagelist()

		parser = argparse.ArgumentParser()
		parser.add_argument("-p", "--page", help="The page to extract: %s or all" % ', '.join(pages.keys()))
		parser.add_argument("-w", "--word", help="The key word(s) to search for.")
		parser.add_argument("-f", "--format", help="The format(s) to search for.")
		parser.add_argument("-t", "--stat", help="The status to search for.")
		parser.add_argument("-s", "--silent", action='store_true', help="If used, no extra parameters will be queried.")
		args = parser.parse_args()
		# print args.echo

		# print "province: " + str(args.province)
		# print "format: " + str(args.format)

		page = args.page
		params = collections.OrderedDict()
		params['srch_word'] = args.word
		params['format'] = args.format
		params['status'] = args.stat
		silent = args.silent

		if page is None:
			answer = raw_input("Please enter the page you would like to use (%s or all): " % ', '.join(pages.keys()))
			if not answer == "":
				page = answer.lower()
			else:
				print "\nERROR: Please specify a web page."
				print "Exiting process."
				sys.exit(1)

		page = page.lower()

		print page

		ext.set_page(page)
		ext.set_params(params)
		ext.run()

	except Exception, err:
		ext.print_log('ERROR: %s\n' % str(err))
		ext.print_log(traceback.format_exc())
		ext.close_log()

		# geoportal_list = extract_geoportal(province)


if __name__ == '__main__':
	sys.exit(main())
