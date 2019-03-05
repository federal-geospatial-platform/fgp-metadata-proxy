import os
import sys
import urllib
import urllib2
from bs4 import BeautifulSoup, Tag, NavigableString, Comment
import collections
import math
import csv
import codecs
import json
import inspect
import requests
import urlparse
import argparse
import traceback
import time
import datetime
import openpyxl

import Main_Extractor as main_ext

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.firefox.options import Options

from common import shared
from common import bsoup
from common import services
#from common import page_group
from common import recurse_ftp as rec_ftp
from common import spreadsheet as sh

class PT_Extractor(main_ext.Extractor):
	def __init__(self):
		''' Initializer for the Extractor class. '''

		# Set the province
		self.province = 'Ontario'
		
		# Initialize the Main Extractor to use its variables
		main_ext.Extractor.__init__(self)
		
		# Create the page groups dictionary
		self.page_groups = []
		
		####################################################################
		# Create Open Data Catalogue page group

		cat_grp = main_ext.PageGroup('opendata', "Ontario Open Data Catalogue")
		
		# Add arguments
		cat_grp.add_arg('word', title='Search Word')
		frm_arg = cat_grp.add_arg('format')
		frm_arg.add_opt('CSV', url_tags=['csv'])
		frm_arg.add_opt('JSON', url_tags=['json'])
		frm_arg.add_opt('KML', url_tags=['kml'])
		frm_arg.add_opt('MDB', url_tags=['mdb'])
		frm_arg.add_opt('TXT', url_tags=['txt'])
		frm_arg.add_opt('XLS', url_tags=['xls'])
		frm_arg.add_opt('XLSX', url_tags=['xlsx'])
		frm_arg.add_opt('ZIP', url_tags=['zip'])
		stat_arg = cat_grp.add_arg('status')
		stat_arg.add_opt('Open', url_tags=['30'])
		stat_arg.add_opt('To be opened', url_tags=['20'])
		stat_arg.add_opt('Under review', url_tags=['10'])
		stat_arg.add_opt('Restricted', url_tags=['0'])
		
		# Add URLs
		cat_grp.add_url('main_url', 'https://www.ontario.ca/search/data-catalogue')
		cat_grp.add_url('api_url', 'https://api.ontario.ca/api/drupal/')
		cat_grp.add_url('srch_url', 'https://api.ontario.ca/es/onesite/_search/template')
		
		# Add to Extractor's page group list
		self.page_groups.append(cat_grp)
		
		
		####################################################################
		# Create Discovering Ontario page group

		disc_grp = main_ext.PageGroup('discovering', "Discovering Ontario")
		
		# Add arguments
		disc_grp.add_arg('word', title='Search Word')
		disc_grp.add_arg('downloadable', title='Downloadable Content')
		sb_arg = disc_grp.add_arg('subpage', debug=True)
		sb_arg.add_opt('Elevation Products', ['elevation'], ['elevation'])
		sb_arg.add_opt('Geology Ontario', ['geology', 'geo'], ['GeologyOntario'])
		sb_arg.add_opt('Geospatial - Non-Commercial Use', ['non-commercial'], ['Non-commercialUse'])
		sb_arg.add_opt('Geospatial - Open Data', ['open'], ['OpenData'])
		sb_arg.add_opt('Geospatial Data', ['geospatial', 'gis'], ['GISData'])
		sb_arg.add_opt('LIO Data Services', ['lio_data'], ['PublicWebServices'])
		sb_arg.add_opt('Land Information Ontario (LIO) Warehouse', ['lio'], ['LIOWarehouse'])
		sb_arg.add_opt('Mapping Applications', ['apps'], ['Application'])
		sb_arg.add_opt('Maps', url_tags=['maps'])
		sb_arg.add_opt('Municipal Planning Provincial Portal (MP3)', ['mp3'], ['MP3'])
		sb_arg.add_opt('Ontario Imagery', ['imagery'], ['Imagery'])
		sb_arg.add_opt('Web Services', ['wms'], ['WMS'])
		
		# Add URLs
		disc_grp.add_url('main_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en')
		#disc_grp.add_url('portal_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en/main.search.embedded')
		#disc_grp.add_url('mdata_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en/metadata.show.embedded')
		#disc_grp.add_url('xml_srch_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en/xml.search')
		#disc_grp.add_url('xml_mdata_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en/xml.metadata.get')
		
		# Add to Extractor's page group list
		self.page_groups.append(disc_grp)
		
		
		####################################################################
		# Create Update Discovering Ontario page group
		
		up_grp = main_ext.PageGroup('update', "Update Discovering Ontario")
		
		# No arguments to add
		
		# Add URLs
		up_grp.add_url('main_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en')
		up_grp.add_url('mdata_url', 'https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en/metadata.show.embedded')
		
		# Add to Extractor's page group list
		self.page_groups.append(up_grp)
		
		
		####################################################################
		# Create Map Services page group
		
		srv_grp = main_ext.PageGroup('services', "Ontario Map Services")
		
		# No arguments to add
		
		# Add URLs
		srv_grp.add_url('geocortex_url', 'https://www.gisapplication.lrc.gov.on.ca/Geocortex/Essentials/essentials42/Rest/sites')
		#srv_grp.add_url('rest_url', 'http://ws.gisdynamic.lrc.gov.on.ca/public/rest/services')
		
		# Add to Extractor's page group list
		self.page_groups.append(srv_grp)
		
		
	###################################################################################################################

	def get_discovering_mdata(self, mdata_xml, url, title_str):
	
		if self.check_result(mdata_xml, url, "Metadata Page - %s" % title_str):
		
			# Get the title
			title_tag = mdata_xml.find('gmd:title')
			title_str = bsoup.get_text(title_tag)
			
			# Get the description
			desc_tag = mdata_xml.find('gmd:abstract')
			desc_str = bsoup.get_text(desc_tag)
			
			# Get the creation date
			recdate_str = ''
			startdate_str = ''
			date_tags = mdata_xml.find_all('gmd:CI_DateTypeCode')
			for datetype in date_tags:
				# print date
				
				datetype_str = datetype['codeListValue']
				
				ci_date = datetype.parent.parent
				
				#print ci_date
				
				date_tag = ci_date.find('gco:Date')
				date_str = bsoup.get_text(date_tag)
				
				if datetype_str == 'creation':
					startdate_str = date_str
				else:
					recdate_str = date_str
			
			# Get the revision date
		
			# Get the metadata standard
			mdata_standard = mdata_xml.find('gmd:metadataStandardName')
			mdata_type = bsoup.get_text(mdata_standard)
			
			# Get the spatial reference
			refsys_tag = mdata_xml.find('gmd:referenceSystemInfo')
			refsys_str = bsoup.get_text(refsys_tag)
			
			# Get the gmd:identificationInfo
			#id_info = mdata_xml.find('gmd:identificationInfo')
			
			# Get the contact
			pub_str = ''
			contact_tag = mdata_xml.find('gmd:contact')
			if contact_tag is not None:
				pub_str = bsoup.get_text(contact_tag.find('gmd:organisationName'))
			
			# Get the data type
			dtype_str = ''
			dist_tag = mdata_xml.find('gmd:distributionFormat')
			if dist_tag is not None:
				dtype_str = bsoup.get_text(dist_tag.find('gmd:name'))
			
			# Get the extents
			extents = ''
			ext_tag = mdata_xml.find('gmd:EX_Extent')
			if ext_tag is not None:
				west = bsoup.get_text(ext_tag.find('gmd:westBoundLongitude'))
				east = bsoup.get_text(ext_tag.find('gmd:eastBoundLongitude'))
				north = bsoup.get_text(ext_tag.find('gmd:northBoundLatitude'))
				south = bsoup.get_text(ext_tag.find('gmd:southBoundLatitude'))
				exts = [north, south, east, west]
			
				extents = shared.create_wkt_extents(exts)
			
			# Get the licensing
			lic_str = ''
			legal_const = mdata_xml.find('gmd:MD_LegalConstraints')
			if legal_const is not None:
				lic_str = bsoup.get_text(legal_const.find('gmd:useLimitation'))
				lic_str = shared.reduce_text(lic_str)
			
			# Get the download(s)
			downloads = []
			transfers = mdata_xml.find_all('gmd:transferOptions')
			for tr in transfers:
				dist_val = bsoup.get_text(tr.find('gmd:unitsOfDistribution'))
				if dist_val == 'Data':
					download_url = bsoup.get_text(tr.find('gmd:URL'))
					downloads.append(download_url)
					
			download_info = shared.get_download_text(download_url=downloads)
			download_str, access_str = download_info.split('|')
			
		else:
			title_str = ''
			desc_str = ''
			startdate_str = ''
			recdate_str = ''
			pub_str = ''
			dtype_str = ''
			lic_str = ''
			extents = ''
			access_str = ''
			download_str = ''
			refsys_str = ''
			mdata_type = ''
			#mdata_url = ''
			
		# FOR DEBUG:
		# if self.debug:
			# print "\n"
			# print "Title: %s" % title_str
			# print "Description: %s" % desc_str
			# print "Publisher: %s" % pub_str
			# print "Type: %s" % dtype_str
			# print "Licensing: %s" % lic_str
			# print "Date: %s" % date_str
			# print "Access: %s" % access
			# print "Download: %s" % download_str
			# print "Spatial Reference: %s" % sp_str
			# print "Metadata Type: %s" % mdata_type
			# print "Metadata URL: %s" % url
			# print "Notes: %s" % self.notes
			# answer = raw_input("Press enter...")

		mdata_items = collections.OrderedDict()
		
		mdata_items['Publisher'] = pub_str
		#pt_csv.add('Data URL'] = '|'.join(data_links)
		mdata_items['Start Date'] = startdate_str
		mdata_items['Recent Date'] = recdate_str
		mdata_items['Extents'] = extents
		mdata_items['Type'] = dtype_str
		mdata_items['Licensing'] = lic_str
		mdata_items['Access'] = access_str
		mdata_items['Download'] = download_str
		mdata_items['Spatial Reference'] = refsys_str
		mdata_items['Metadata Type'] = mdata_type
		mdata_items['Notes'] = self.notes
		
		return mdata_items
				
			# # Get date from adjacent element of <th> with text 'Date'
			# date_str = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Date', url=url)

			# # Get description from adjacent element of <th> with text 'Abstract'
			# desc_str = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Abstract', url=url)
			# #print desc_str

			# # Get data type from adjacent element of <th> with text 'Environment description'
			# dtype_str = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Environment description', url=url)

			# # Get the publisher from adjacent element of <th> with text 'Organisation name'
			# pub_str = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Organisation name', url=url)

			# # Get Data URL(s)
			# download_items = mdata_soup.find_all('span', string='Data for download')
			# data_links = []
			# for i in download_items:
				# th_html = i.parent
				# td_html = th_html.find_next_sibling('td')
				# data_link = td_html.a['href']
				# data_links.append(data_link)

			# # Get metadata type from adjacent element of <th> with text 'Metadata standard name'
			# mdata_type = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Metadata standard name', url=url)

			# # Get licence from adjacent element of <th> with text 'Access constraints'
			# #lic_str = bsoup.get_adj_text_by_label(mdata_soup, 'th', 'Access constraints')
			# tags = [('th', 'Use limitation'), ('th', 'Other constraints'), ('th', 'Access constraints')]
			# lic_str = self.get_mdata(mdata_soup, tags)
			# #print mdata_url
			# #print "lic_str: %s" % lic_str

			# # Get the spatial reference
			# ref_span = mdata_soup.find('span', string='Reference System Information')
			# sp_str = ''
			# if ref_span is not None:
				# ref_td = bsoup.get_parent(ref_span, 'td')
				# sp_str = bsoup.get_adj_text_by_label(ref_td, 'th', 'Code', url=url)

			# # Fill in the access and download_str variables based on the number of
			# #   data_links
			# if len(data_links) == 0:
				# download_str = 'No'
				# access = 'Contact the Province'
			# elif len(data_links) == 1:
				# download_str = data_links[0]
				# access = 'Download/Web Accessible'
			# else:
				# download_str = 'Multiple Downloads'
				# access = 'Download/Web Accessible'
		
	def get_category(self, cat_div):
		# Get the onclick text to get the category
		cat_text = cat_div['onclick']
		#print "cat_text: %s" % cat_text
		start_pos = cat_text.find("(")
		end_pos = cat_text.find(")", start_pos)
		category = cat_text[start_pos+2:end_pos-1]
		#print "category: %s" % category
		
		return category
		
	def get_categories(self, main_url):
		# Get a list of categories from the main page
		attrb = ['id', 'latest_updates']
		main_soup = bsoup.get_soup(main_url, True, attrb)
		
		print "\nMain URL: %s" % main_url
		
		if not self.check_result(main_soup, main_url, "Ontario Discover Portal"):
			return None

		divs = main_soup.find_all('div')

		# Find all <div> within <div> with class 'geosearchfields' to get the different categories
		cat_divs = []
		for div in divs:
			if div.has_attr('class'):
				#print div['class']
				if div['class'][0].find('geosearchfields') > -1:
					cat_divs = div.find_all('div')
					break
					
		# Convert the category list of elements to words
		categories = [self.get_category(c) for c in cat_divs]
		
		return categories
		
	def get_cat_mdata(self, json, tag):
		mdatags = json['metatags']
		
		if not tag in mdatags: return ''
		mdata_tag = mdatags[tag]
		attached = mdata_tag['#attached']
		drupal = attached['drupal_add_html_head']
		
		sub_drupal = drupal[0][0]
		
		#print "sub_drupal: %s" % sub_drupal
		
		mdata_val = sub_drupal['#value']
		
		return mdata_val
		
	def get_hits(self, url, params):
		''' Get the hits from the catalogue
		'''
		
		params['size'] = 10
		payload = {"params": params, 
					"template":
					{
						"id": "dataset"
					}
				}
		
		json_payload = json.dumps(payload)
		print json_payload
		
		# Get the results to get the total record count
		result = requests.post(url, data=json_payload)
		json_res = result.json()
		
		#print json_res
		
		# Get the total number of records
		total_hits = json_res['hits']['total']
		params['size'] = total_hits
		payload = {"params": params, 
					"template":
					{
						"id": "dataset"
					}
				}
				
		print payload
		
		# Rerun POST
		result = requests.post(url, data=json.dumps(payload))
		json_res = result.json()
		
		hits = json_res['hits']['hits']
		
		return hits
	
	def get_next_twenty(self, driver):
		''' Gets the element with text 'Next 20'
		:param driver: The Selenium driver
		:return: Element with text 'Next 20'
		'''
		found_elements = driver.find_elements_by_class_name("secondary")
		for el in found_elements:
			if el.text.strip() == 'Next 20':
				return el
				
	def get_mdata(self, mdata_soup, heading_tups):
		
		mdata_val = ''
		for heading in heading_tups:
			if mdata_val == '':
				tag = heading[0]
				name = heading[1]
				mdata_val = bsoup.get_adj_text_by_label(mdata_soup, tag, name)
				
		return mdata_val

	def extract_opendata(self): #, word=None, format=None, status=None):
		''' Extracts the results from the Open Data Catalogue of Ontario
		:param word: The word used to filter the results
		:param format: The format used to filter the results
		:param status: The status used to filter the results
		:return: None
		'''
		
		# Get the parameters
		word = self.get_arg_val('word')
		format = self.get_arg_val('format')
		status = self.get_arg_val('status')

		# URL examples for statuses:
		# https://www.ontario.ca/search/data-catalogue?sort=asc&status=%5B%2230%22%5D
		# https://www.ontario.ca/search/data-catalogue?sort=asc&status=["30"]

		# Full example of possible URL:
		# https://www.ontario.ca/search/data-catalogue?
		# sort=asc
		# &query=gis
		# &filetype=%5B%22csv%22%5D
		# &publisher=%5B%22Advanced%20Education%20and%20Skills%20Development%22%5D
		# &topic=%5B%22Arts%20and%20culture%22%5D
		# &status=%5B%2230%22%5D

		# Different statuses:
		# 30 - Open
		# 20 - To be opened
		# 10 - Under review
		# 0 - Restricted

		main_url = self.pg_grp.get_url('main_url')
		api_url = self.pg_grp.get_url('api_url')
		srch_url = self.pg_grp.get_url('srch_url')

		self.print_log("\nExtracting from %s" % self.pg_grp.get_title())
		
		self.print_title("Extracting Ontario's Open Data Catalogue")
		
		start_time = datetime.datetime.now()
		print "Process started at: %s" % start_time

		# Create the CSV file
		csv_fn = "Catalogue_results"
		pt_csv = sh.PT_CSV(csv_fn, self)
		pt_csv.open_csv()

		# Set the parameters for the URL query
		params = collections.OrderedDict()
		params['lang'] = 'en'
		params['from'] = 0
		#params['size'] = 10
		if word is not None and not word == '':
			params['query'] = word
		if format is not None and not format == '':
			params['fileType?'] = {'list': ['%s' % format]}
		if status is not None and not status == '':
			stat_tag = self.get_arg('status').get_urltags()[0]
			status_val = {'list': [stat_tag]}
			# if status.lower() == 'open':
				# status_val = {'list': ['30']}
			# elif status.lower() == 'to be opened':
				# status_val = {'list': ['20']}
			# elif status.lower() == 'under review':
				# status_val = {'list': ['10']}
			# else:
				# status_val = {'list': ['0']}
			params['status?'] = status_val
			
		results = shared.get_hits(srch_url, params)
		
		print "Number of results: %s" % len(results)
		
		for idx, res in enumerate(results):
			try:
				msg = "Extracting record %s of %s" % (idx + 1, len(results))
				shared.print_oneliner(msg)
			
				# Get the title
				title_str = res['_source']['title']
				
				# Get the URL for the API
				res_url = res['_source']['url']
				mdata_url = 'https://www.ontario.ca/%s' % res_url
				
				# Get the API results of the dataset
				api_ds_url = '%s%s' % (api_url, res_url.replace('/', '%2F'))
				api_json = shared.get_json(api_ds_url)
				
				# Get the description
				desc_str = self.get_cat_mdata(api_json, 'description')
				
				# Get the subject for the keywords
				keywords = self.get_cat_mdata(api_json, 'dcterms:subject')
				
				# Get the subject for the keywords
				publish_str = self.get_cat_mdata(api_json, 'dcterms:publisher')
				
				#print api_json.keys()
				
				# Get the external file links
				downloads = []
				if 'field_od_external_file_link' in api_json:
					link_tag = api_json['field_od_external_file_link']
					ext_links = link_tag['und']
					
					for l in ext_links:
						download = l['url']
						if download.find('http') == -1:
							download = 'https:%s' % download
						downloads.append(download)
						
				# Get other download info
				if 'field_odr_file' in api_json:
					odr_tag = api_json['field_odr_file']
					files = odr_tag['und']
					for f in files:
						if 'uri' in f:
							download = f['uri']
						else:
							download = f['url']
						if download.find('http') == -1:
							download = 'https:%s' % download
						downloads.append(download)
						
				# Get the formats from the downloads
				formats = []
				for d in downloads:
					d_split = d.split('.')
					format = d_split[len(d_split) - 1]
					formats.append(format.upper())
				
				download_info = shared.get_download_text(formats, downloads)
				download_str, access_str = download_info.split('|')
				
				formats = list(set(formats))
				
				# Get the recent date
				recdate_str = self.get_cat_mdata(api_json, 'article:modified_time')
				if recdate_str.find('T') > -1:
					recdate_str = recdate_str.split('T')[0]
				
				# Get the licensing
				licensing = ''
				if 'field_intellectual_property_url' in api_json:
					property_tag = api_json['field_intellectual_property_url']
					licensing = property_tag['und'][0]['title']
					licensing = shared.filter_unicode(licensing)
					
				# Get the coverage
				coverage = ''
				if 'field_od_geographical_coverage' in api_json:
					cover_tag = api_json['field_od_geographical_coverage']
					coverage = cover_tag['und'][0]['safe_value']
				
				# Get the start date
				startdate_str = ''
				if 'field_od_time_period_start' in api_json:
					period_tag = api_json['field_od_time_period_start']
					und_val = period_tag['und'][0]
					if und_val is not None:
						startdate_str = und_val['value']
				
				pt_csv.add('Source', "Ontario Data Catalogue")
				pt_csv.add('Title', title_str)
				pt_csv.add('Description', desc_str)
				pt_csv.add('Start Date', startdate_str)
				pt_csv.add('Recent Date', recdate_str)
				pt_csv.add('Publisher', publish_str)
				pt_csv.add('Keywords', keywords)
				pt_csv.add('Access', access_str)
				pt_csv.add('Available Formats', '|'.join(formats))
				pt_csv.add('Download', download_str)
				pt_csv.add('Licensing', licensing)
				pt_csv.add('Metadata URL', mdata_url)
				pt_csv.add('Notes', self.notes)

				# Write the results to the CSV
				pt_csv.write_dataset()
			except:
				print
				print api_ds_url
				print traceback.print_exc(file=sys.stdout)
				answer = raw_input("Press enter...")

		# # Build the URL query
		# query_url = shared.get_post_query(main_url, params)

		# print "\nQuery URL: %s" % query_url

		# # Create a Selenium driver with a headless Firefox
		# #options = Options()
		# #options.add_argument("--headless")
		# driver = webdriver.Firefox()#firefox_options=options)
		# driver.implicitly_wait(20)
		# driver.get(query_url)

		# # Wait for the element with ID "rechercheEdelForm:lancerRecherche" to load
		# shared.wait_page_load(driver, By.CLASS_NAME, "results-page")

		# # Continually click 'Next 20' button until it no longer exists
		# next_20 = self.get_next_twenty(driver)
		# while next_20 is not None:
			# next_20.click()
			# driver.implicitly_wait(20)
			# next_20 = self.get_next_twenty(driver)

		# # Grab the results of the results page once all results are shown
		# page_html = driver.page_source

		# # Create the BeautifulSoup object
		# soup = BeautifulSoup(page_html, 'html.parser')

		# # Get all the datasets on the current page (all datasets are in a 'div' with class 'dataset-item')
		# page_list = soup.find_all('ul', attrs={'class': 'results-page'})

		# # Get the number of records for status purposes
		# res_span = bsoup.find_tags_containing(soup, 'results', 'span')
		# res_text = res_span.text
		# res_text = res_text.strip()
		# end_pos = res_text.find(' results')
		# num_results = res_text[:end_pos]

		# print "The number of results is %s" % num_results

		# record_count = 0

		# # Cycle through each page DIV
		# for page in page_list:

			# li_res = page.find_all('li')

			# for li in li_res:

				# record_count += 1

				# #print "Extracting record %s of %s" % (record_count, num_results)
				# msg = "Extracting record %s of %s" % (record_count, num_results)
				# shared.print_oneliner(msg)

				# # Get the title and URL
				# h3 = li.find('h3')
				# a = h3.find('a')
				# title_str = a.contents[0].strip()
				# mdata_href = a['href']
				# mdata_url = urlparse.urljoin(query_url, mdata_href)

				# date_str = ''
				# pub_str = ''
				# dt_list = li.find_all('dt')
				# for dt in dt_list:
					# # Get the date
					# if dt.text.find("Date added") > -1:
						# dd_html = dt.find_next_sibling('dd')
						# date_str = dd_html.string.strip()

					# # Get the publisher
					# elif dt.text.find("Publisher") > -1:
						# dd_html = dt.find_next_sibling('dd')
						# pub_str = dd_html.text.strip()

				# # Get the rest of the values using the metadata URL
				# attrb = ('class', 'thumbs-down')
				# mdata_soup = bsoup.get_soup(mdata_url, True, attrb)
				
				# if self.check_result(mdata_soup, mdata_url, 
					# "Metadata Page - %s" % title_str, output=False):

					# # Get the description
					# div_pgbody = mdata_soup.find('div', attrs={'id': 'pagebody'})
					# desc_str = bsoup.get_text(div_pgbody)

					# # Get the downloads
					# h2 = bsoup.get_adj_tags_by_text(mdata_soup, 'h2', 'Download data', True)
					# if len(h2) == 0:
						# download_str = 'No'
						# access = 'Contact the Province'
						# format_list = []
					# else:
						# dwnload_div = h2[0].parent
						# anchors = dwnload_div.find_all('a')
						# if len(anchors) == 0:
							# download_str = 'No'
							# access = 'Contact the Province'
						# elif len(anchors) == 1:
							# download_str = anchors[0]['href']
							# access = "Download/Web Accessible"
						# else:
							# download_str = 'Multiple Downloads'
							# access = "Download/Web Accessible"

						# # Get the formats from the downloads
						# format_list = []
						# for a in anchors:
							# format = a.abbr.text
							# format_list.append(format)
							
				# else:
					# # Set the variables as empty strings if the metadata can't be loaded
					# desc_str = ''
					# date_str = ''
					# pub_str = ''
					# access = ''
					# download_str = ''
					# format_list = []
					# mdata_url = ''

				# # Add all values to the CSV file
				# pt_csv.add('Title', title_str)
				# pt_csv.add('Description', desc_str)
				# pt_csv.add('Date', date_str)
				# pt_csv.add('Publisher', pub_str)
				# pt_csv.add('Access', access)
				# pt_csv.add('Download', download_str)
				# #pt_csv.add('Licensing', lic_str)
				# pt_csv.add('Available Formats', '|'.join(format_list))
				# pt_csv.add('Metadata URL', mdata_url)
				# pt_csv.add('Notes', self.notes)

				# # Write the results to the CSV
				# pt_csv.write_dataset()
				
				# self.notes = ''

		# driver.quit()

		pt_csv.close_csv()
		
		# Print ending time
		end_time = datetime.datetime.now()
		print "\nExtraction complete at %s." % end_time
		tot_time = end_time - start_time
		print "It took %s to complete." % tot_time

	def extract_discovering(self): #, word=None, downloadable=False):
		''' Extracts results from the Land Information Ontario geoportal
		:param word: The word used to filter the results
		:param downloadable: Determines whether to include results only with downloads
		:return: None
		'''
		
		self.print_log("\nExtracting from %s" % self.pg_grp.get_title())
		
		self.print_title("Extracting Discovering Ontario Data")
		
		start_time = datetime.datetime.now()
		print "Process started at: %s" % start_time
		
		# Here are the different options available in the GeoNetwork
		# https://www.javacoeapp.lrc.gov.on.ca/geonetwork/srv/en
		#	Home Page: /main.home
		#	Search Page: /main.search.embedded
		#	Metadata Page: /metadata.show.embedded
		#	XML Search: /xml.search
		#	XML Metadata Page: /xml.metadata.get
		
		word = self.get_arg_val('word')
		downloadable = self.get_arg_val('downloadable')
		subpage = self.get_arg_val('subpage')
		
		# Get all URLs
		srv_url = self.pg_grp.get_url('main_url')
		#portal_url = self.pg_grp.get_url('portal_url')
		#mdata_query_url = self.pg_grp.get_url('mdata_url')
		#xml_srch_url = self.pg_grp.get_url('xml_srch_url')
		#xml_mdata_url = self.pg_grp.get_url('xml_mdata_url')
		
		main_page = '%s/main.home' % srv_url
		categories = self.get_categories(main_page)
		
		for category in categories:
			
			params = collections.OrderedDict()
			params['any'] = word
			params['category'] = category
			params['download'] = downloadable
		
			# Create the XML search for the category
			#cat_url = '%s/xml.search?category=%s' % (srv_url, category)
			srch_url = '%s/xml.search' % srv_url
			cat_url = shared.get_post_query(srch_url, params)
			
			#print "cat_url: %s" % cat_url
			#answer = raw_input("Press enter...")
			
			csv_fn = "%s_results" % category.title()
			pt_csv = sh.PT_CSV(csv_fn, self)
			pt_csv.open_csv()
			
			print "Getting XML '%s'..." % cat_url
			cat_xml = bsoup.get_xml_soup(cat_url)
			
			results = cat_xml.find_all('metadata')
			
			print "\nNumber of results: %s" % len(results)
			
			for idx, res in enumerate(results):
				msg = "Extracting %s of %s records for category '%s'" % \
						(idx + 1, len(results), category)
				shared.print_oneliner(msg)
				
				# Get the ID of the dataset
				ds_id = bsoup.get_text(res.find('id'))
				
				# Get the title
				title_str = bsoup.get_text(res.find('title_eng'))
				
				# Get the description
				desc_str = bsoup.get_text(res.find('abstract_eng'))
				
				# Get the keywords
				kywrd_tags = res.find_all('keyword')
				keywords = [bsoup.get_text(k) for k in kywrd_tags]
				
				# # Get the start date
				# startdate_str = bsoup.get_text(res.find('createDate'))
				# if startdate_str.find('T') > -1:
					# startdate_str = startdate_str.split('T')[0]
				
				# # Get the recent date
				# recdate_str = bsoup.get_text(res.find('changeDate'))
				# if recdate_str.find('T') > -1:
					# recdate_str = recdate_str.split('T')[0]
				
				# Build the metadata HTML page
				mdata_url = '%s/metadata.show.embedded?id=%s' % (srv_url, ds_id)
				
				# Build the metadata XML to get the rest of the information
				mdata_xml_url = '%s/xml.metadata.get?id=%s' % (srv_url, ds_id)
				#mdata_xml_json = bsoup.xml_to_dict(mdata_xml_url)
				mdata_xml = bsoup.get_xml_soup(mdata_xml_url)
				
				mdata_info = self.get_discovering_mdata(mdata_xml, mdata_xml_url, title_str)
				
				for k, v in mdata_info.items():
					pt_csv.add(k, v)
					
				pt_csv.add('Source', 'Discovering Ontario Data')
				pt_csv.add('Title', title_str)
				pt_csv.add('Description', desc_str)
				pt_csv.add('Keywords', ', '.join(keywords))
				
				# Write the results to the CSV
				pt_csv.write_dataset()
				
			print
			
		# Print ending time
		end_time = datetime.datetime.now()
		print "\nExtraction complete at %s." % end_time
		tot_time = end_time - start_time
		print "It took %s to complete." % tot_time
		
		# # Get the parameters
		# word = self.argmt['word'].get_value()
		# subpage = self.argmt['subpage'].get_value()
		# downloadable = self.argmt['downloadable'].get_value()

		# self.print_log("\nExtracting from %s" % self.pg_grp.get_title())
		
		# self.print_title("Extracting Discovering Ontario Data")
		
		# start_time = datetime.datetime.now()
		# print "Process started at: %s" % start_time

		# main_url = self.pg_grp.get_url('main_url')
		# portal_url = self.pg_grp.get_url('portal_url')
		# mdata_query_url = self.pg_grp.get_url('mdata_url')

		# # Load the main page to get a list of categories
		# attrb = ['id', 'latest_updates']
		# main_soup = bsoup.get_soup(main_url, True, attrb)
		
		# print "\nMain URL: %s" % main_url
		
		# if not self.check_result(main_soup, main_url, "Ontario Discover Portal"): return None

		# divs = main_soup.find_all('div')

		# # Find all <div> within <div> with class 'geosearchfields' to get the different categories
		# cat_divs = []
		# for div in divs:
			# if div.has_attr('class'):
				# #print div['class']
				# if div['class'][0].find('geosearchfields') > -1:
					# cat_divs = div.find_all('div')
					# break
					
		# # Convert the category list of elements to words
		# categories = [self.get_category(c) for c in cat_divs]
		
		# print "subpage: %s" % subpage
		# print "categories: %s" % categories
		# # If the subpage is in the available categories,
		# #	set the list of categories to the subpage
		# for c in categories:
			# if subpage.lower() == c.lower():
				# categories = [c]
				# break

		# for category in categories:
			
			# # Create the CSV file
			# csv_fn = "%s_results" % category.title()
			# pt_csv = sh.PT_CSV(csv_fn, self)
			# #if pt_csv.check_exists():
			# #	print "CSV file for category '%s' already exists." \
			# #		  " Proceeding to next category." % category
			# pt_csv.open_csv()

			# # Set the parameters for the URL query
			# params = collections.OrderedDict()
			# if word is not None and not word == "": params['any'] = word
			# params['category'] = category
			# if downloadable: params['download'] = 'on'
			# params['hitsPerPage'] = '1000'

			# # Build the URL query
			# query_url = shared.get_post_query(portal_url, params)

			# print "\nQuery URL: %s" % query_url

			# # Get the soup of the query URL
			# xml_soup = bsoup.get_xml_soup(query_url)
			
			# if not self.check_result(xml_soup, query_url, "Ontario Discover Portal Query"): continue

			# items = xml_soup.find_all('div', attrs={'class': 'hit'})

			# num_items = len(items)

			# print "\nNumber of records: " + str(num_items)

			# record_count = 0

			# for index, item in enumerate(items):
			
				# record_count += 1
				# msg = "Extracting %s of %s records for category '%s'" % (record_count, num_items, category)
				# shared.print_oneliner(msg)

				# # Get the ID from <div> with class 'thumbnail_results'
				# thumb_div = item.find('div', attrs={'class': 'thumbnail_results'})
				# rating_link = thumb_div.a['id']
				# id_str = rating_link.replace("rating.link.", "")
				# #print "ID: %s" % id_str

				# # Get the title from <div> with class 'hittitle'
				# title_str = item.find('div', attrs={'class': 'hittitle'}).text
				
				# #answer = raw_input("Press enter...")

				# # Metadata URL:
				# mdata_url = '%s?id=%s' % (mdata_query_url, id_str)
				
				# print mdata_url
				# answer = raw_input("Press enter...")

				# # HTML Access
				# mdata_soup = bsoup.get_soup(mdata_url) #, True, ('class', 'padded-content'))
				
				# mdata_items = self.get_discovering_mdata(mdata_soup, mdata_url, title_str)
				
				# for k, v in mdata_items.items():
					# pt_csv.add(k, v)

				# # Write the results to the CSV
				# pt_csv.write_dataset()

			# print

		# pt_csv.close_csv()
		
		# # Print ending time
		# end_time = datetime.datetime.now()
		# print "\nExtraction complete at %s." % end_time
		# tot_time = end_time - start_time
		# print "It took %s to complete." % tot_time
		
	def update_discover(self):
		''' Extracts the latest Discovering Ontario information
		
		'''
		
		self.print_title("Updating Discovering Ontario Data")
		
		start_time = datetime.datetime.now()
		print "Process started at: %s" % start_time
		
		excel_url = 'https://www.sse.gov.on.ca/sites/MNR-PublicDocs/EN/CMID/DataDistributionCatalogue.xlsx'
		
		mdata_query_url = self.pg_grp.get_url('mdata_url')
		srv_url = self.pg_grp.get_url('main_url')

		now = datetime.datetime.now()
		date_str = now.strftime('%Y%m%d')

		# Download the XLSX file
		xlsx_fn = 'files\\DataDistributionCatalogue_Update_%s.xlsx' % date_str
		xlsx_f = urllib.urlretrieve(excel_url, xlsx_fn)

		#r = requests.get(excel_url)  # make an HTTP request

		wb = openpyxl.load_workbook(xlsx_fn)

		sheets = wb.sheetnames
		
		url_list = []

		for sh_idx, sh_name in enumerate(sheets):
			print "\nSheet %s:" % sh_name

			sheet = wb[sh_name]
			
			row_count = sheet.max_row
			
			if sh_idx == 0:
				row_start = 5
			else:
				row_start = 1
				
			for row in range(row_start, row_count - row_start + 1):
				cell = sheet['A%s' % row]
				hp_link = cell.hyperlink
				if hp_link is not None:
					#print hp_link
					title_str = cell.value
					url = hp_link.target
					if url is not None:
						print url
						
						# Parse the URL to get the ID
						id_str = url.split('=')[1]
						
						# Metadata URL:
						#mdata_url = '%s?uuid=%s' % (mdata_query_url, id_str)
						mdata_url = '%s/xml.metadata.get?uuid=%s' % (srv_url, id_str)
						
						#print mdata_url
						#answer = raw_input("Press enter...")
						
						url_list.append((title_str, mdata_url))
						
				#answer = raw_input("Press enter...")
						
		#answer = raw_input("Press enter...")
						
		# Create the CSV file
		csv_fn = "Discovering_Update_%s_results" % date_str
		pt_csv = sh.PT_CSV(csv_fn, self)
		pt_csv.open_csv()
		
		for idx, info in enumerate(url_list):
			msg = "Extracting %s of %s records" % \
				(idx + 1, len(url_list))
			shared.print_oneliner(msg)
			
			title_str, url = info
		
			mdata_soup = bsoup.get_xml_soup(url)
			
			if not self.check_result(mdata_soup, url, "Discovering Update"):
				continue
			
			mdata_items = self.get_discovering_mdata(mdata_soup, url, title_str)

			for k, v in mdata_items.items():
				pt_csv.add(k, v)
				
			pt_csv.add('Source', 'Discovering Ontario Data')
			pt_csv.add('Title', title_str)
			pt_csv.add('Description', desc_str)

			# Write the results to the CSV
			pt_csv.write_dataset()

		print

		pt_csv.close_csv()
		
		# Print ending time
		end_time = datetime.datetime.now()
		print "\nExtraction complete at %s." % end_time
		tot_time = end_time - start_time
		print "It took %s to complete." % tot_time
		
	def extract_services(self):
		
		# Create the CSV file for the province
		csv_fn = "Services_results"
		pt_csv = sh.PT_CSV(csv_fn, self)
		pt_csv.open_csv()
		
		geocortex_url = self.pg_grp.get_url('geocortex_url')
		
		geocortex = services.PT_Geocortex(geocortex_url)
		
		site_data = geocortex.get_layers()
		
		if not self.check_result(site_data, geocortex_url, 'GeoCortex Map'): return None
					
		#filter_rows = shared.process_duplicates(site_data)
		
		for index, rec in enumerate(site_data):
			shared.print_oneliner("Adding %s of %s to CSV inventory" \
									% (index + 1, len(site_data)))
			for k, v in rec.items():
				pt_csv.add(k, v)
			
			pt_csv.add('Source', 'BC Map Services')
			pt_csv.write_dataset()
		
		print

		#print site_data
		
		#answer = raw_input("Press enter...")

def main():
	# parser.add_argument("-t", "--tool", help="The tool to use: %s" % ', '.join(tool_list))
	# parser.add_argument("-w", "--word", help="The key word(s) to search for.")
	# parser.add_argument("-f", "--format", help="The format(s) to search for.")
	# parser.add_argument("-c", "--category", help="The category(ies) to search for.")
	# parser.add_argument("-d", "--downloadable", help="Determines wheter to get only downloadable datasets.")
	# parser.add_argument("-l", "--html", help="The HTML file to scrape (only for OpenData website).")
	# parser.add_argument("-s", "--silent", action='store_true', help="If used, no extra parameters will be queried.")

	ext = Extractor()

	try:
		pages = ext.get_pagelist()

		parser = argparse.ArgumentParser()
		parser.add_argument("-p", "--page", help="The page to extract: %s or all" % ', '.join(pages.keys()))
		parser.add_argument("-w", "--word", help="The key word(s) to search for.")
		parser.add_argument("-f", "--format", help="The format(s) to search for.")
		parser.add_argument("-t", "--stat", help="The status to search for.")
		parser.add_argument("-s", "--silent", action='store_true', help="If used, no extra parameters will be queried.")
		args = parser.parse_args()
		# print args.echo

		# print "province: " + str(args.province)
		# print "format: " + str(args.format)

		page = args.page
		params = collections.OrderedDict()
		params['srch_word'] = args.word
		params['format'] = args.format
		params['status'] = args.stat
		silent = args.silent

		if page is None:
			answer = raw_input("Please enter the page you would like to use (%s or all): " % ', '.join(pages.keys()))
			if not answer == "":
				page = answer.lower()
			else:
				print "\nERROR: Please specify a web page."
				print "Exiting process."
				sys.exit(1)

		page = page.lower()

		print page

		ext.set_page(page)
		ext.set_params(params)
		ext.run()

	except Exception, err:
		ext.print_log('ERROR: %s\n' % str(err))
		ext.print_log(traceback.format_exc())
		ext.close_log()

		# geoportal_list = extract_geoportal(province)


if __name__ == '__main__':
	sys.exit(main())
