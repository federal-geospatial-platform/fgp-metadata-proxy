import codecs
import json
import os
import sys
import shutil
import urllib2
from bs4 import BeautifulSoup
import collections
from urlparse import urlparse
import csv
import inspect
import traceback
import argparse
import glob
from openpyxl import *
from openpyxl.styles import *
from openpyxl.worksheet.write_only import WriteOnlyCell
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from operator import itemgetter

import shared

class PT_CSV:
	def __init__(self, fn, pt_ext=None, header=None, out_folder=None, title=None, 
					read_only=False):
					
		if not read_only and pt_ext is None:
			print "\nWARNING: A province/territory Extractor must be specified " \
					"to create a CSV file."
			return None
					
		self.fn = fn
		self.pt_ext = pt_ext
		self.read_only = read_only
		
		self.work_folder = os.path.dirname(
							os.path.dirname(
							os.path.dirname(
							os.path.realpath(__file__))))
		
		if self.pt_ext is not None:
			self.province = self.pt_ext.get_province()
			self.res_folder = os.path.join(os.sep, self.work_folder, 'results', 
										self.province)
		else:
			self.province = None
			self.res_folder = None
		
		self.csv_file = None
		
		if os.path.exists(self.fn) and self.read_only:
			self.csv_path = self.fn
		else:
			self.csv_path = None
		self.out_folder = out_folder

		# Set the headers of the CSV file if not specified
		if header is None:
			# self.header = ['Source', 'Title', 'Description', 'Type', 'Start Date', 
							# 'Recent Date', 'Update Frequency', 
							# 'Publisher', 'Licensing', 'Available Formats', 
							# 'Access', 'Download', 'Spatial Reference', 
							# 'Data URL', 'Web Page URL', 'Web Map URL', 
							# 'Service', 'Service Name', 'Service URL', 
							# 'Metadata URL', 'Metadata Type', 'Notes']
			self.header = get_header_info()['csv']
		else:
			self.header = header

		# Create the data dictionary
		self.data_dict = collections.OrderedDict((k, "") for k in self.header)

		self.datasets = None
		self.title = title

		# Create the CSV file
		if not self.read_only or self.csv_path is None:
			self.create_csv_fn()

	def add(self, param, value):
		''' Add a dataset to the data dictionary.
		:param param: The name of the parameter (key).
		:param value: The value of the parameter.
		:return: None
		'''

		self.data_dict[param] = value

	def add_header_item(self, hd_item):
		''' Adds header (columns) to the top of the CSV file
		:param hd_item: A list of header values.
		:return: None
		'''

		self.header.append(hd_item)
		self.data_dict = collections.OrderedDict((k, "") for k in self.header)

	def check_exists(self):
		''' Checks to see if the CSV file exists.
		:return: True is it exists, False if it does not.
		'''

		if os.path.exists(self.csv_path):
			return True
		else:
			return False

	def close_csv(self):
		''' Close the CSV file.
		:return: None
		'''

		print "\nClosing CSV file '%s'." % self.csv_path
		self.csv_file.close()

	def create_csv_fn(self):
		''' Creates the CSV file name
		:return: None
		'''
		
		# Check for debug
		debug_str = ''
		#print dir(self.pt_ext)
		if self.pt_ext is not None:
			if self.pt_ext.debug:
				debug_str = '_debug'

		# Get the CSV filename
		if self.province is None:
			# If no province is provided, set the output name to only include 
			#	self.fn
			self.csv_path = os.path.join(os.sep, self.work_folder, 
										"%s%s.csv" % (self.fn, debug_str))
		else:
			if self.out_folder is None:
				# If no output folder is provided, set the output folder to 
				#	'<work_folder>/<province>/results'
				self.out_folder = os.path.join(os.sep, self.work_folder, 
												'results', str(self.province))
			if not os.path.isdir(self.out_folder):
				# If the out_folder is not a folder, set the output folder 
				#	to '<work_folder>/<province>/results'
				self.out_folder = os.path.join(os.sep, self.work_folder, 
												'results', self.province)

			# Create the CSV filename
			self.csv_path = os.path.join(os.sep, self.out_folder, 
										"%s_%s%s.csv" % (self.province, 
														self.fn, debug_str))

	def get_dictrows(self):
		''' Gets a list of rows as dictionaries.
		'''
		
		if not self.read_only: return None
		
		# Grab each line in the CSV file
		csv_lines = self.csv_file.readlines()
		
		# Get the header
		self.header = shared.get_header(csv_lines)
		
		row_lst = []
		
		for idx, line in enumerate(csv_lines[1:]):
			msg = "Extracting %s of %s lines" % (idx + 1, len(csv_lines[1:]))
			shared.print_oneliner(msg)
			
			# Convert the line to a dictionary
			row_dict = self.row_to_dict(line, self.header)
			
			if row_dict is None: continue
			
			#if row_dict['Source'] == '':
			#	row_dict = source
			
			# Store the title and description in a list
			row_lst.append(row_dict)
			
		print
		
		return row_lst
														
	def get_header(self):
		''' Get the header (columns) of the CSV file.
		:return: A list of header names.
		'''

		return self.header

	def open_csv(self, f_mode='w'):
		''' Opens the CSV file based on the specified mode.
		:param f_mode: The mode used when opening the file.
		:return: None
		'''

		# Create the filename
		if self.read_only:
			f_mode = 'r'
			add_header = False
		else:
			self.create_csv_fn()

			# Check to see if the file exists for adding the header
			add_header = True
			if os.path.exists(self.csv_path) and f_mode == 'a':
				add_header = False

		# Open the document
		print "\nCSV File: " + str(self.csv_path)
		try:
			self.csv_file = codecs.open(self.csv_path, encoding='utf_8_sig', 
									mode=f_mode)
		except IOError as e:
			print "\nERROR: %s" % e
			print "If the CSV file '%s' is opened, please close it." \
					% self.csv_path
			print "Closing extraction process."
			sys.exit(1)

		if add_header:
			# Write the title in the CSV file
			if self.title is not None:
				self.write_line("Web Page Title: " + self.title)
				
			#print "self.header: %s" % self.header
			#answer = raw_input("Press enter...")

			# Write the header in the CSV file
			if self.header is not None:
				self.write_header(self.header)

	def remove_duplicates(self, unique_field, url=False):
		''' Removes duplicate entries in the CSV file based on the 
				unique_field as the column.'''

		return None
		
	def row_to_dict(self, row, header):
	
		#print header
		#answer = raw_input("Press enter...")

		# Convert each row in the row list to a CSV reader object
		try:
			filtered_row = shared.filter_unicode(row, french=True)
			filtered_row = filtered_row.replace('\r', ' ')
			#csv_reader = sh.UnicodeReader(StringIO(filtered_row))
			
			entry_lst = shared.parse_csv_row(filtered_row)
			
			#entry_lst = [item for item in csv_reader]
		except Exception, e:
			# If an error occurs, filter the row to remove unicode characters
			filtered_row = shared.filter_unicode(row, french=True)
			#csv_reader = csv.reader(StringIO(shared.filter_unicode(filtered_row)))
			
			entry_lst = shared.parse_csv_row(filtered_row)
			
		
		if len(entry_lst) == 0:
			print "\n"
			print "\nWARNING: The current row is empty."
			answer = raw_input("Press enter...")
			return None
				
		if not len(entry_lst) == len(header):
			print "\n"
			print
			print "Header: %s" % header
			print
			print "Input Row: %s" % row
			print
			print "Filtered Row: %s" % filtered_row
			print
			print "Entry List: %s" % entry_lst
			print
			print len(entry_lst)
			print "\nWARNING: The current row does not contain the correct " \
					"number of columns."
			answer = raw_input("Press enter...")
		
		# Convert entry to dictionary
		entry_dict = collections.OrderedDict()
		for idx, h in enumerate(header):
			entry_dict[h] = entry_lst[idx]
			
		return entry_dict

	def write_dataset(self, dataset=None):
		''' Writes the specified or current dataset to the CSV file.
		:param dataset: A dataset to add to the CSV file.
		:return: None
		'''
		try:
			if dataset is None: dataset = self.data_dict
			# values = [(k, v) for k, v in dataset.items()]
			str_vals = []
			for k, v in dataset.items():
				v = shared.filter_unicode(v)
				#v = clean_text(v)
				if v is not None:
					v = v.replace('"', '""')
					v = v.replace('","', '", "')
					v = '"%s"' % v.strip().replace("\n", " ").replace("\r", " ")
					v = shared.clean_text(v)
				else:
					v = '""'
				str_vals.append(v)
				
			#if any(s in 'Building Outlines' for s in str_vals):
			#print str_vals
			#answer = raw_input("Press enter...")

			# Check to see if all values are blank and if so,
			#   don't add to the CSV file
			empty_check = [v for v in str_vals if not v.strip() == '']
			
			#print "empty_check: %s" % empty_check

			if len(empty_check) > 0:
				self.csv_file.write(','.join(str_vals) + "\n")
				
			#answer = raw_input("Press enter...")

			# Clear the dataset values
			self.data_dict = collections.OrderedDict(
								(k, "") for k in self.header)
			
			self.pt_ext.set_notes('')
		except:
			print "Exception in user code:"
			print '-' * 60
			if 'k' in vars(): print k
			traceback.print_exc(file=sys.stdout)
			print '-' * 60
			self.pt_ext.set_notes('')
			for k, v in dataset.items():
				print "%s: %s" % (k, v)
			answer = raw_input("Press enter...")

	def write_datasets(self, ds_list):
		''' Adds a list of datasets to the CSV file
		:param ds_list: A list of dataset information.
		:return: None
		'''

		for ds in ds_list:
			self.write_dataset(ds)

	def write_header(self, fieldnames=None):
		''' Write the header to the top of the CSV file.
		:param fieldnames: A list of fieldnames
		:return: None
		'''

		if fieldnames is not None: self.header = fieldnames

		if self.header is None:
			print "\nWARNING: Cannot write header to CSV. No header provided."
			return None
		self.csv_file.write(','.join(self.header) + "\n")

	def write_line(self, line):
		''' Write a line to the CSV file.
		:param line: A string of the line to put into the CSV file.
		:return: None
		'''
		self.csv_file.write(line + "\n")

	def write_list(self, item_list):
		''' Write a list of items to a single line in the CSV file.
		:param item_list: A list of items.
		:return: None
		'''
		self.csv_file.write(','.join(item_list) + "\n")

	def write_url(self, url=None, url_name=None):
		''' Writes a URL to the top of the CSV file.
		:param url: The URL of the page.
		:param url_name: The name of the page.
		:return: None
		'''

		if url is not None: self.url = url
		if url_name is not None: self.url_name = url_name

		if self.url is None:
			print "\nWARNING: Cannot write URL to CSV. No URL provided."
			return None

		if self.url_name is None:
			self.csv_file.write("Main URL:,%s\n" % self.url)
		else:
			self.csv_file.write("%s:,%s\n" % (self.url_name, self.url))
		
		
class PT_XL:
	
	def __init__(self, pt_ext=None, fn=None, header=[], out_folder=None, 
				read_only=False, write_only=False, setup=True, replace=False, 
				replace_ws=False, silent=True, ws_title=None, fn_word=None):
				
		self.wb = None
		self.ws = None
		#self.fn_base = os.path.basename(self.fn)
		self.row = collections.OrderedDict()
		self.read_only = read_only
		self.write_only = write_only
		self.ws_headers = collections.OrderedDict()
		#self.out_folder = out_folder
			
		if pt_ext is None and fn is None:
			print "\nWARNING: A filename or a PT_Extractor object must be " \
					"specified to create a PT_XL object."
			return None
			
		if pt_ext is not None:
			# If a PT_Extractor object was provided
			self.pt_ext = pt_ext
			# Set the province
			self.province = self.pt_ext.get_province()
			# Set the working folder
			self.work_folder = os.path.dirname(
							os.path.dirname(
							os.path.dirname(
							os.path.realpath(__file__))))
			# Set the results folder
			self.res_folder = os.path.join(os.sep, self.work_folder, 'results', 
										self.province)
										
			if out_folder is None:
				self.out_folder = self.res_folder
			else:
				self.out_folder = out_folder
			
			# Get the current page_group of the province/territory
			self.pg_grp = self.pt_ext.get_pg_grp()
			# Set the current page_group ID
			self.pg_id = self.pg_grp.get_id()
			self.pg_title = self.pg_grp.get_title()
			if ws_title is None: ws_title = self.pg_id
			
			# Create the sheet name with the page group and search word, 
			#	if applicable
			srch_word = self.pt_ext.get_arg_val('word')
			if srch_word is None or srch_word == '':
				srch_word = ''
			else:
				if ws_title is not None and not ws_title == '':
					srch_word = ' - %s' % srch_word
			self.ws_name = "%s%s" % (ws_title, srch_word)
			
			# Set the output Excel filename
			debug_str = ''
			if self.pt_ext.debug:
				debug_str = '_debug'
			
			if fn_word is None or fn_word == '':
				fn = '%s_%s%s_results.xlsx' % (self.province, self.pg_id, \
												debug_str)
			else:
				fn = '%s_%s_%s%s_results.xlsx' % (self.province, self.pg_id, \
												fn_word, debug_str)
			self.xl_fn = os.path.join(os.sep, self.out_folder, fn)
			
		else:
			# If only a filename was specified
			
			# Set the output folder
			self.xl_fn = fn
			self.out_folder = os.path.dirname(self.xl_fn)
			
			# Set the sheetname
			self.ws_name = ws_title
			
			# Set the ID
			self.pg_id = ws_title
			
		# Check if Excel file is available for editing
		if os.path.exists(self.xl_fn):
			try:
				os.rename(self.xl_fn, self.xl_fn)
			except Exception, e:
				print e
				print "The file '%s' cannot be accessed. " \
						"Please make sure the file is not opened." % self.xl_fn
				sys.exit(1)

		# Set the headers of the CSV file if not specified
		if len(header) == 0:
			# self.header = [('Source', 50), ('Title', 100), ('Description', 100), 
						# ('Type', 20), ('Start Date', 20), ('Recent Date', 20), 
						# ('Update Frequency', 35), ('Publisher', 60), 
						# ('Licensing', 50), ('Available Formats', 50), 
						# ('Access', 32), ('Download', 25), 
						# ('Spatial Reference', 50), ('Data URL', 70), 
						# ('Web Page URL', 70), ('Web Map URL', 70), 
						# ('Service', 25), ('Service Name', 100), 
						# ('Service URL', 70), ('Metadata URL', 70), 
						# ('Metadata Type', 65), ('Notes', 100)]
			self.header = get_header_info()['xl']
			self.header_txt = [h[0] for h in self.header]
			self.widths = [h[1] for h in self.header]
		else:
			self.header = header
			if isinstance(header[0], tuple):
				self.header_txt = [h[0] for h in self.header]
				self.widths = [h[1] for h in self.header]
			else:
				self.header_txt = self.header
				self.widths = []
				
		#print "header: %s" % self.header_txt
		#print "widths: %s" % self.widths
		
		#answer = raw_input("Press enter...")
		
		self.header_cells = []
		
		self.silent = silent
		self.replace_ws = replace_ws
		
		# Remove exiting Excel file if applicable
		if os.path.exists(self.xl_fn) and replace:
			os.remove(self.xl_fn)
		
		# Don't setup anything if read-only
		if self.read_only: setup = False
		
		#print "Setup: %s" % setup
		
		#answer = raw_input("Press enter...")
		
		if setup:
			# Setup every if specified
		
			# Setup workbook
			if self.wb is None:
				self.create_workbook()
				
			# Setup worksheet
			if self.ws is None and self.ws_name is not None:
				self.add_worksheet(self.ws_name)
				
				# Setup header
				if len(self.header) > 0:
					#print "self.ws: %s" % self.ws
					#answer = raw_input("Press enter...")
					self.add_header()
		
	def add_cell(self, cell_val, column=None, ws_name=None, 
				**kwargs):
		''' Adds a cell to the current row list (self.row).
		:param cell_val: The value for the cell.
		:param fill: The openpyxl.styles.fills object for the cell.
		:param font: The openpyxl.styles.fonts object for the cell.
		:param border: The openpyxl.styles.borders object for the cell.
		:param alignment: The openpyxl.styles.alignment object for the cell.
		:param number_format: The openpyxl.styles.numbers object for the cell.
		:param protection: The openpyxl.styles.protection object for the cell.
		'''
		
		fill = None
		font = None
		border = None
		alignment = None
		number_format = None
		protection = None
		
		if kwargs is not None:
			if 'fill' in kwargs: fill = kwargs['fill']
			if 'font' in kwargs: font = kwargs['font']
			if 'border' in kwargs: border = kwargs['border']
			if 'alignment' in kwargs: alignment = kwargs['alignment']
			if 'number_format' in kwargs:
				number_format = kwargs['number_format']
			if 'protection' in kwargs: protection = kwargs['protection']
		
		if self.read_only:
			print "\nADD_CELL WARNING: Cannot write cell to '%s'. " \
					"It is set to read-only." % self.pg_id
			return None
		
		if self.ws is None:
			print "\nADD_CELL WARNING: Cannot write cell to '%s'. " \
					"No worksheet set." % self.pg_id
			print "Please use 'add_worksheet' or 'set_worksheet' method " \
					"in the PT_XL to set the worksheet."
			return None
			
		# Set the worksheet if specified, if not use the current worksheet
		if ws_name is not None:
			wrksheet = self.wb[ws_name]
		else:
			wrksheet = self.ws
			
		#if col_idx is not None:
		#	new_cell = Cell(wrksheet, value=cell_val, row=1)
		#else:
		#print "cell_val: %s" % cell_val
		new_cell = WriteOnlyCell(wrksheet, value=cell_val)
		
		if fill is not None: new_cell.fill = fill
		if font is not None: new_cell.font = font
		if border is not None: new_cell.border = border
		if alignment is not None: new_cell.alignment = alignment
		if number_format is not None: new_cell.number_format = number_format
		if protection is not None: new_cell.protection = protection
		#if cell_idx is not None: new_cell.cell_idx
		#self.row.append((col_idx, new_cell))
		if isinstance(column, int):
			header_vals = self.ws_header[self.ws.title]
			column = header_vals[column]
		
		#print "column: %s" % column
		
		if column is None:
			row_lst = []
			if 'row_lst' in self.row:
				row_lst = self.row['row_lst']
			row_lst.append(new_cell)
			self.row['row_lst'] = row_lst
		else:
			self.row[column] = new_cell
		
	def write_cell(self, col, row, value):
		''' Writes a single sell to a specific location in the sheet.
		'''
		self.ws.cell(column=col, row=row, value=value)

	def add_header(self, in_header=None, ws_name=None):
		''' Add the given header (either in_header or self.header) 
			to the current Worksheet.
		'''
		
		if self.ws is None and ws_name is None:
			print "\nADD_HEADER WARNING: Cannot add header to '%s'. " \
					"No worksheet set." % self.pg_id
			print "Please use 'add_worksheet' or 'set_worksheet' method " \
					"in the PT_XL to set the worksheet."
			return None
		
		if self.read_only:
			print "\nADD_HEADER WARNING: Cannot write header to '%s'. " \
					"It is set to read-only." % self.pg_id
			return None
		
		if in_header is None and len(self.header) == 0:
			print "\nADD_HEADER WARNING: No header has been set. Please " \
					"enter a in_header value for the 'add_header' method."
			return None
				
		if in_header is not None: self.header = in_header
		
		if len(self.header) == 0: self.header = in_header
		
		# Set the worksheet if specified, if not use the current worksheet
		if ws_name is not None:
			wrksheet = self.wb[ws_name]
		else:
			wrksheet = self.ws
		
		# Get a list of the widths from the header list
		if isinstance(self.header[0], tuple):
			self.header_txt = [h[0] for h in self.header]
			self.widths = [h[1] for h in self.header]
		else:
			self.header_txt = self.header
			self.widths = [100 for h in self.header]
			
		self.ws_headers[wrksheet.title] = self.header_txt
		
		# Set the widths to 100 characters
		self.set_column_widths(self.widths)
	
		self.header_cells = []
		for h in self.header_txt:
			#print "Worksheet: %s" % self.ws
			#print "Value: %s" % h
			cell = WriteOnlyCell(wrksheet, value=h)
			cell.fill = PatternFill("solid", fgColor="C6EFCE")
			cell.font = Font(name='Calibri (Body)', size=14, color="006100")
			cell.alignment = Alignment(horizontal='center')
			thin = Side(border_style="thin", color="A6A6A6")
			cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
			self.header_cells.append(cell)
			
		wrksheet.append(self.header_cells)
		
		return self.header_cells
		
	def add_item(self, column, item, ws_name=None, **kwargs):
		''' Adds an item to the Excel spreadsheet
		'''
		
		self.add_cell(item, column, ws_name, **kwargs)
		
	def add_title(self, title, merge=(1, 2)):
		''' Adds a title to the next row in the current worksheet.
		'''
		
		start_col, end_col = merge
		self.merge_cells(start_col, end_col)
		
		fill = PatternFill("solid", fgColor="C6EFCE")
		font = Font(name='Calibri (Body)', size=14, color="006100")
		thin = Side(border_style="thin", color="A6A6A6")
		border = Border(top=thin, left=thin, right=thin, bottom=thin)
		self.add_cell(title, 0, fill=fill, font=font, border=border)
		self.write_row()
		
	def add_worksheet(self, ws_name='', header=None, title=None,
						replace=None):
		''' Add a worksheet to the current Workbook (self.wb).
		'''
		if self.read_only:
			print "\nADD_WORKSHEET WARNING: Cannot add worksheet to '%s'. " \
				"It is set to read-only."
			return None
			
		if self.wb is None:
			print "\nADD_WORKSHEET WARNING: Cannot add worksheet to '%s'. " \
					"No Workbook set." % self.pg_id
			print "Please use 'create_workbook' or 'set_workbook' method " \
					"in the PT_XL to set the Workbook."
			return None
			
		if replace is None:
			replace = self.replace_ws
			
		if len(ws_name) > 30:
			ws_name = ws_name[:30]
		
		if ws_name in self.wb.sheetnames:
			if replace:
				self.wb.remove(self.wb[ws_name])
				self.ws = self.wb.create_sheet(ws_name)
			else:
				self.ws = self.wb[ws_name]
		else:
			print "Worksheet name: %s" % ws_name
			self.ws = self.wb.create_sheet(ws_name)
			
		# Set the widths to 100 characters
		if header is not None:
			widths = [h[1] for h in header]
			self.set_column_widths(widths)
		
		if title is not None:
			self.add_title(title)
			
		if header is not None:
			self.add_header(header, ws_name)
			
		return self.ws
		
	def close_workbook(self):
		''' Closes the current workbook.
		'''
		
		print "Closing workbook."
		self.wb.close()
		
	def create_workbook(self):
		''' Creates a new Workbook.
		'''
		
		if self.read_only:
			if not self.silent:
				print "\nCREATE_WORKBOOK WARNING: Cannot create Workbook to " \
					"'%s'. It is set to read-only." % self.pg_id
				print "Open existing Workbook for read-only."
			if os.path.exists(self.xl_fn):
				self.wb = load_workbook(self.xl_fn)
			else:
				print "\nCREATE_WORKBOOK WARNING: '%s' does not exist. " \
						"Cannot open Workbook." % self.xl_fn
			return None
			
		if os.path.exists(self.xl_fn):
			print "\nNOTE: The Workbook '%s' already exists. " \
					"Opening it for edit." % self.xl_fn
			self.wb = load_workbook(self.xl_fn)
		else:
			self.wb = Workbook(self.write_only)
			
	def delete_sheet(self, ws_name):
		''' Deletes the specified worksheet.
		'''
		if ws_name in self.wb.sheetnames:
			del self.wb[ws_name]

	def get_abbreviation(self, pt):
		''' Conver the province/territory name to its 2-letter abbreviation.
		'''
		
		if pt == 'Alberta':
			return 'AB'
		elif pt == 'British Columbia':
			return 'BC'
		elif pt == 'Manitoba':
			return 'MB'
		elif pt == 'New Brunswick':
			return 'NB'
		elif pt == 'Newfoundland & Labrador':
			return 'NL'
		elif pt == 'Nova Scotia':
			return 'NS'
		elif pt == 'Northwest Territories' or pt == 'NWT':
			return 'NT'
		elif pt == 'Nunavut':
			return 'NU'
		elif pt == 'Ontario':
			return 'ON'
		elif pt == 'Prince Edward Island' or pt == 'PEI':
			return 'PE'
		elif pt == 'Quebec':
			return 'QC'
		elif pt == 'Saskatchewan':
			return 'SK'
		elif pt == 'Yukon':
			return 'YT'
		else:
			return pt
			
	# def get_column_by_name(self, col_name):
		# ''' Gets the column index based on a header name.
		# '''
		
		# # Get the header row
		# rows = self.ws.rows
		# row = rows.next()
		
		# print row
		
		# answer = raw_input("Press enter...")
	
	def get_header(self):
		''' Gets the header (top row) from the Excel spreadsheet.
		'''
		
		if self.ws.title in self.ws_headers.keys():
			return self.ws_headers[self.ws.title]
		
		# Get the first row
		rows = self.ws.rows
		row = rows.next()
		
		self.header_txt = []
		
		for c in row:
			#print c.value
			self.header_txt.append(c.value)
			
		#print self.header_txt
			
		self.ws_headers[self.ws.title] = self.header_txt
		
	def get_rows(self):
		''' Gets all the rows for the current Worksheet.
		'''
		
		if self.ws is None:
			print "\nGET_ROWS WARNING: Cannot get rows from '%s'. " \
					"No worksheet set." % self.pg_id
			print "Please use 'add_worksheet' or 'set_worksheet' method " \
					"in the PT_XL to set the worksheet."
			return []
		
		return self.ws.rows
		
	def get_dictrows(self, output='cells'):
		''' Gets a list of rows as dictionaries.
		'''
		
		#for row in self.ws.rows:
		
		if not self.ws.title in self.ws_headers.keys():
			return None
		
		header_vals = self.ws_headers[self.ws.title]
		
		print header_vals
			
		rows = []
		for idx, row in enumerate(self.ws.rows):
			if idx == 0: continue
			
			#print len(header_vals)
			#print len(row)
			
			row_dict = collections.OrderedDict()
			for idx, h in enumerate(header_vals):
				if output == 'values':
					if row[idx].value is None:
						row_dict[h] = ''
					else:
						row_dict[h] = row[idx].value
				else:	
					row_dict[h] = row[idx]
			rows.append(row_dict)
			
			#print rows
		
			#answer = raw_input("Press enter...")
		
		return rows
		
	def set_column_widths(self, widths):
		''' Sets the columns widths of the current sheet
		'''
		
		print "Setting widths..."
		
		#print widths
		#answer = raw_input("Press enter...")
	
		for i, w in enumerate(widths):
			#last_letter = chr(i + 97).upper()
			last_letter = get_column_letter(i + 1)
			self.ws.column_dimensions[last_letter].width = w
			
	def set_workbook(self, fn="", read_only=True):
		''' Loads an existing Workbook (Excel file)
		'''
		if fn == '':
			fn = self.xl_fn
		
		self.wb = load_workbook(fn, read_only)
			
	def save_file(self):
		''' Saves the current Workbook (self.wb).
		'''
		
		if self.wb is None:
			print "\nSAVE_FILE WARNING: Cannot save file '%s'. " \
					"No Workbook set." % self.pg_id
			print "Please use 'create_workbook' or 'set_workbook' method " \
					"in the PT_XL to set the Workbook."
			return None
			
		self.delete_sheet('Sheet')
		
		self.wb.save(self.xl_fn)
		
	def get_worksheet(self):
		''' Gets the current worksheet of the Workbook.
		'''
		return self.ws
		
	def get_ws_list(self):
		''' Gets a list of worksheet names.
		'''
		
		ws_names = self.wb.sheetnames
		
		ws_list = []
		for ws in ws_names:
			ws_list.append(self.wb[ws])
			
		return ws_list
		
	def get_ws_name(self):
		''' Gets the name of the current worksheet.
		'''
		
		return self.ws.title
	
	def set_worksheet(self, ws_name=None, ws=None):
		''' Set the current Worksheet (self.ws) to a specified 
			existing worksheet.
			NOTE: To create a new worksheet, use 'add_worksheet'.
		'''
	
		if ws is None and ws_name is None:
		#	print "\nSET_WORKSHEET WARNING: No worksheet info given. " \
		#		"Please enter either the sheet name or the worksheet object."
			self.ws = self.wb.active
		
		if ws is not None:
			self.ws = ws
			
		if self.wb is None:
			print "\nSET_WORKSHEET WARNING: Cannot save file '%s'. " \
					"No Workbook set." % self.pg_id
			print "Please use 'create_workbook' or 'set_workbook' method " \
					"in the PT_XL to set the Workbook."
			return None
			
		if ws_name is not None:
		
			if len(ws_name) > 30:
				ws_name = ws_name[:30]
		
			if isinstance(ws_name, int) or ws_name.isdigit():
				ws_lst = self.wb.sheetnames
				ws_name = ws_lst[int(ws_name)]
				self.ws = self.wb[ws_name]
			elif ws_name in self.wb.sheetnames:
				self.ws = self.wb[ws_name]
			else:
				print "\nSET_WORKSHEET WARNING: Sheet '%s' does not " \
						"exist in '%s'." % (ws_name, self.pg_id)
				return None
				
		self.get_header()
				
	def style_range(self, cell_range, border=Border(), fill=None, 
					font=None, alignment=None):
		"""
		Apply styles to a range of cells as if they were a single cell.

		:param range: An excel range to style (e.g. A1:F20)
		:param border: An openpyxl Border
		:param fill: An openpyxl PatternFill or GradientFill
		:param font: An openpyxl Font object
		"""

		top = Border(top=border.top)
		left = Border(left=border.left)
		right = Border(right=border.right)
		bottom = Border(bottom=border.bottom)

		first_cell = self.ws[cell_range.split(":")[0]]
		if alignment:
			self.ws.merge_cells(cell_range)
			first_cell.alignment = alignment

		rows = self.ws[cell_range]
		if font:
			first_cell.font = font

		for cell in rows[0]:
			cell.border = cell.border + top
		for cell in rows[-1]:
			cell.border = cell.border + bottom

		for row in rows:
			l = row[0]
			r = row[-1]
			l.border = l.border + left
			r.border = r.border + right
			if fill:
				for c in row:
					c.fill = fill

				
	def merge_cells(self, start_col, end_col):
		''' Merges cells in the row after the current row.
		'''
		
		cur_row = self.ws._current_row + 1
		
		self.ws.merge_cells(start_column=start_col, start_row=cur_row, \
							end_column=end_col, end_row=cur_row)
		
		#answer = raw_input("Press enter...")
		
	def move_sheet(self, ws_name, pos=None):
		''' Moves the position of a sheet's tab
			:param ws_name: The name of the sheet to move.
			:param pos: The final position of the sheet 
						(if None, the sheet is placed at the end).
		'''
		
		wrksheet = self.wb[ws_name]
		
		# Get the position of the specified sheet
		ws_index = self.wb.index(wrksheet)
		
		if pos is None:
			self.wb._sheets.append(self.wb._sheets.pop(ws_index))
		else:
			self.wb._sheets.insert(pos, self.wb._sheets.pop(ws_index))
				
	def write_list(self, row_list=[], ws_name=None):
		''' Writes a list as a row.
		'''
		
		if ws_name is not None:
			self.ws = self.wb[ws_name]
		
		for idx, col in enumerate(row_list):
			#print "worksheet name: %s" % ws_name
			self.add_cell(col, idx)
			
		self.write_row()
			
	def write_row(self, ws_name=None):
		''' Write the current row to the current Worksheet.
		'''
		
		if self.read_only:
			print "\nWRITE_ROW WARNING: Cannot write row to '%s'. " \
					"It is set to read-only." % self.pg_id
			return None
			
		if self.ws is None:
			print "\nWRITE_ROW WARNING: Cannot write row to '%s'. " \
					"No worksheet set." % self.pg_id
			print "Please use 'add_worksheet' or 'set_worksheet' method " \
					"in the PT_XL to set the worksheet."
			return None
		
		# # Add self.row to the worksheet
		
		# # Rearrange row into the proper order
		# print "\nUnsorted:"
		# for i, r in self.row:
			# print '%s: %s' % (i, r.value)
		
		# out_row = self.row
		# if len(self.row) > 0:
			# check_tup = self.row[0]
		
			# if isinstance(check_tup, tuple):
				# max_itm = max([r[0] for r in self.row])
				
				# out_row = []
				# for i in range(max_itm):
				# out_row = [r[1] for r in sort_row]
		
		# #print "\nSorted:"
		# #print ['%s: %s' % (i, r.value) for i, r in sort_row]
		# #print "Out row:"
		# #print ['%s' % r.value for r in out_row]
		
		# Switch to the specified worksheet, otherwise write
		#	to the current sheet
		if ws_name is not None:
			self.ws = self.wb[ws_name]
		
		out_row = []
		
		if len(self.row.keys()) == 0:
			self.ws.append(out_row)
			self.row = collections.OrderedDict()
			return None
		
		if not self.ws.title in self.ws_headers.keys():
			# If the current worksheet has no header text
			#	in self.ws_headers
			out_row = self.row['row_lst']
		else:
			# If the current sheet has header in self.ws_headers
			header_vals = self.ws_headers[self.ws.title]
			for h in header_vals:
				if h in self.row.keys():
					out_row.append(self.row[h])
				else:
					out_row.append('')
		
		# print
		# print len(out_row)
		# print out_row
		# answer = raw_input("Press enter...")
				
		self.ws.append(out_row)
		
		self.row = collections.OrderedDict()
		
	def ws_exists(self, ws_name):
		''' Checks to see if a worksheet exists.
		'''
		
		if ws_name in self.wb.sheetnames:
			return True
			
		return False
		
class UTF8Recoder:
	"""
	Iterator that reads an encoded stream and reencodes the input to UTF-8
	"""

	def __init__(self, f, encoding):
		self.reader = codecs.getreader(encoding)(f)

	def __iter__(self):
		return self

	def next(self):
		return self.reader.next().encode("utf-8")


class UnicodeReader:
	"""
	A CSV reader which will iterate over lines in the CSV file "f",
	which is encoded in the given encoding.
	"""

	def __init__(self, f, dialect=csv.excel, encoding="utf-8", **kwds):
		f = UTF8Recoder(f, encoding)
		self.reader = csv.reader(f, dialect=dialect, **kwds)

	def next(self):
		row = self.reader.next()
		return [unicode(s, "utf-8") for s in row]

	def __iter__(self):
		return self


class UnicodeWriter:
	"""
	A CSV writer which will write rows to CSV file "f",
	which is encoded in the given encoding.
	"""

	def __init__(self, f, dialect=csv.excel, encoding="utf-8", **kwds):
		# Redirect output to a queue
		self.queue = cStringIO.StringIO()
		self.writer = csv.writer(self.queue, dialect=dialect, **kwds)
		self.stream = f
		self.encoder = codecs.getincrementalencoder(encoding)()

	def writerow(self, row):
		self.writer.writerow([s.encode("utf-8") for s in row])
		# Fetch UTF-8 output from the queue ...
		data = self.queue.getvalue()
		data = data.decode("utf-8")
		# ... and reencode it into the target encoding
		data = self.encoder.encode(data)
		# write to the target stream
		self.stream.write(data)
		# empty queue
		self.queue.truncate(0)

	def writerows(self, rows):
		for row in rows:
			self.writerow(row)

		
def unicode_csv_reader(utf8_data, dialect=csv.excel, **kwargs):

    csv_reader = csv.reader(utf8_data, dialect=dialect, **kwargs)
    for row in csv_reader:
        yield [unicode(cell, 'utf-8') for cell in row]
		
		
def get_header_info(out_type='pt'):
	''' This method contains the different header information.
	'''
	
	out_header = collections.OrderedDict()
	
	base_header = [('Source', 30), ('Title', 100), ('Description', 100), 
					('Type', 20), ('Start Date', 20), ('Recent Date', 20), 
					('Update Frequency', 35), ('Publisher', 60), 
					('Licensing', 50), ('Keywords', 50), 
					('Available Formats', 50), 
					('Access', 32), ('Download', 25), 
					('Spatial Reference', 50), ('Extents', 40), 
					('Data URL', 70), ('Web Page URL', 70), 
					('Web Map URL', 70), ('Service', 25), 
					('Service Name', 100), ('Service URL', 70), 
					('Metadata URL', 70), ('Metadata Type', 65), 
					('Notes', 100)]
	csv_base = [h[0] for h in base_header]
	
	if out_type == 'analysis':
		#out_header['csv'] = ['Layer', 'P/T'] + csv_base
		xl_header = [('Word Found', 20), ('Found In', 15), ('Layer', 24)] + \
						base_header
	elif out_type == 'ca_both':
		xl_header = base_header[:3] + [('FGP Topic Category', 30), 
					('FGP Metadata URL', 70), ('FGP Metadata Standard', 70), 
					('FGP Metadata Keywords', 50), ('Open Maps Topic Category', 30), 
					('Open Maps Metadata URL', 70), ('Open Maps Metadata Standard', 70), 
					('Open Maps Metadata Keywords', 50)]
	elif out_type == 'ca_fgp':
		xl_header = base_header[:3] + [('Topic Category', 30), 
					('Subject Thesaurus', 30), 
					('Metadata ID', 40), ('Metadata Keywords', 50), 
					('Metadata URL', 70), ('Metadata XML URL', 70)]
	elif out_type == 'ca_portal':
		xl_header = base_header[:3] + [base_header[19]]
	elif out_type == 'shp':
		xl_header = [('Source', 30), ('Title', 100), ('Descript', 100), 
					('Type', 20), ('Start', 20), ('Recent', 20), 
					('UpDt_Freq', 35), ('Publisher', 60), 
					('Licensing', 50), ('Keywords', 50), 
					('Formats', 50), ('Access', 32), 
					('Download', 25), ('Sp_Ref', 50), 
					('Extents', 100), ('Data_URL', 70), 
					('Page_URL', 70), ('Map_URL', 70), 
					('Service', 25), ('Serv_Name', 100), 
					('Serv_URL', 70), ('Mdata_URL', 70), 
					('MdataType', 65), ('Notes', 100)]	
	else:
		xl_header = base_header
		
	out_header['xl'] = xl_header
	out_header['csv'] = [h[0] for h in xl_header]
		
	return out_header
	
	