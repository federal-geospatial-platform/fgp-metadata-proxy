import os
import sys
import urllib2
from bs4 import BeautifulSoup, Comment
import collections
import math
import csv
import re
import numpy as np
import json
import urlparse
import argparse
import traceback
import datetime
import inspect
import time
import pprint
import codecs
from openpyxl import *
from openpyxl.styles import *
from openpyxl.worksheet.write_only import WriteOnlyCell
from StringIO import StringIO
from pyPdf import PdfFileWriter, PdfFileReader

import Main_Extractor as main_ext

from operator import itemgetter
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.firefox.options import Options

from common import shared
from common import bsoup
#from common import access_rest as rest
#from common import page_group
from common import recurse_ftp as rec_ftp
from common import spreadsheet as sh

class PT_Extractor(main_ext.Extractor):
	def __init__(self):
		# Set the province
		self.province = 'Canada'
		
		# Initialize the Main Extractor to use its variables
		main_ext.Extractor.__init__(self)
		
		# Set the Organization options
		self.org_opts = []
		self.org_opts.append(['Agriculture and Agri-Food Canada', \
							['aafc'], ['aafc-aac']])
		self.org_opts.append(['Canadian Northern Economic Development Agency', \
							['cannor'], ['cannor']])
		self.org_opts.append(['Elections Canada', \
							['elections'], ['elections']])
		self.org_opts.append([\
						'Environment and Climate Change Canada', \
						['eccc'], ['ec']])
		self.org_opts.append(['Fisheries and Oceans Canada', \
						['dfo'], ['dfo-mpo']])
		self.org_opts.append([\
						'Indigenous and Northern Affairs Canada', \
						['inac'], ['aandc-aadnc']])
		self.org_opts.append(['Natural Resources Canada', \
						['nrcan'], ['nrcan-rncan']])
		self.org_opts.append(['Parks Canada', \
						['pc'], ['pc']])
		self.org_opts.append(['Statistics Canada', \
						['statcan'], ['statcan']])
		self.org_opts.append(['Transport Canada', \
						['tc'], ['tc']])
		self.org_opts.append([\
						'Canadian Environmental Assessment Agency', \
						['ceaa'], ['ceaa-acee']])
		self.org_opts.append([\
						'Canadian Food Inspection Agency', \
						['cfia'], ['cfia-acia']])
		self.org_opts.append(['Canadian Space Agency', \
						['csa'], ['csa-asc']])
		self.org_opts.append([\
						'Crown-Indigenous Relations and Northern Affairs', \
							['aandc'], ['aandc-aadnc']])
		self.org_opts.append(['Health Canada', \
						['hc'], ['hc-sc']])
		self.org_opts.append(['Veterans Affairs Canada', \
						['vac'], ['vac-acc']])
		
		# Create the page groups dictionary
		self.page_groups = []
		
		####################################################################
		# Create Open Government page group

		open_grp = main_ext.PageGroup('portal', 'Open Government Portal')
		
		# Add arguments
		open_grp.add_arg('word', title='Search Word')
		ds_type = open_grp.add_arg('ds_type', title='Portal Type')
		ds_type.add_opt('Open Data', ['dataset'], ['dataset'])
		ds_type.add_opt('Open Information', ['info'], ['info'])
		open_grp.add_arg('category', default='fgp', title='Collection/Category')
		open_grp.add_arg('keyword')
		org_arg = open_grp.add_arg('organization')
		for org in self.org_opts:
			org_arg.add_opt(org[0], org[1], org[2])
		
		# Add URLs
		open_grp.add_url('main_url', 'https://open.canada.ca/data/en/dataset')
		
		# Add to Extractor's page group list
		self.page_groups.append(open_grp)
		
		
		####################################################################
		# Create FGP page group
		
		fgp_grp = main_ext.PageGroup('fgp', 'Federal Geospatial Platform')
		
		# Add arguments
		fgp_grp.add_arg('word', title='Search Word')
		fgp_grp.add_arg('start', debug=True, title='Starting Record')
		fgp_grp.add_arg('test', debug=True, title='Test Site')
		org_arg = fgp_grp.add_arg('organization')
		for org in self.org_opts:
			org_arg.add_opt(org[0], org[1], org[2])
		
		# Add URLs
		fgp_grp.add_url('main_url', 'https://gcgeo.gc.ca/geonetwork/search/eng')
		fgp_grp.add_url('query_url', 'https://gcgeo.gc.ca/geonetwork/srv/eng/xml.search')
		fgp_grp.add_url('csw_url', 'https://gcgeo.gc.ca/geonetwork/srv/eng/csw')
		fgp_grp.add_url('mdata_url', 'https://gcgeo.gc.ca/geonetwork/srv/eng/xml.metadata.get')
		
		# Add to Extractor's page group list
		self.page_groups.append(fgp_grp)
		
		
		####################################################################
		# Create All Platforms page group
		
		all_grp = main_ext.PageGroup('all', 'All Platforms')
		
		# Add arguments
		all_grp.add_arg('word', title='Search Word')
		all_grp.add_arg('category', default='fgp', title='Collection/Category')
		org_arg = all_grp.add_arg('organization')
		for org in self.org_opts:
			org_arg.add_opt(org[0], org[1], org[2])
		
		# Add URLs
		all_grp.add_url('fgp_main_url', 'https://gcgeo.gc.ca/geonetwork/search/eng')
		all_grp.add_url('fgp_query_url', 'https://gcgeo.gc.ca/geonetwork/srv/eng/rest.search')
		all_grp.add_url('fgp_mdata_url', 'https://gcgeo.gc.ca/geonetwork/srv/eng/xml.metadata.get')
		all_grp.add_url('open_url', 'https://open.canada.ca/data/en/dataset')
		
		# Add to Extractor's page group list
		self.page_groups.append(all_grp)
		
		
	def get_province(self):
		''' Gets the province name of the extractor.
		:return: The province name of the extractor.
		'''
		
		return self.province

	###################################################################################################################

	def compare(self, fgp_rows, open_rows):
	
		if (fgp_rows is None or len(fgp_rows) == 0) and \
			(open_rows is None or len(open_rows) == 0):
			final_duplicates = []
			unique_fgp = []
			unique_open = []
			return unique_fgp, unique_open, final_duplicates
	
		if fgp_rows is None or len(fgp_rows) == 0:
			final_duplicates = []
			unique_fgp = []
			unique_open = open_rows
			return unique_fgp, unique_open, final_duplicates
		
		if open_rows is None or len(open_rows) == 0:
			final_duplicates = []
			unique_fgp = fgp_rows
			unique_open = []
			return unique_fgp, unique_open, final_duplicates
		
		# Get a list of duplicate indices from both sites
		#duplicates = []
		dup_indices = []
		for f_idx, f_row in enumerate(fgp_rows):
			f_title = f_row['Title']
			
			for o_idx, o_row in enumerate(open_rows):
				o_title = o_row['Title']
				if f_title == o_title:
					#duplicates.append((f_row, o_row))
					dup_indices.append((f_idx, o_idx))
					
		# Get the unique rows and duplicate rows
		duplicates = []
		unique_fgp = []
		unique_open = []
		
		for idx, f in enumerate(fgp_rows):
			fgp_indices = [i[0] for i in dup_indices]
			#print fgp_indices
			if idx in fgp_indices:
				i = fgp_indices.index(idx)
				open_idx = dup_indices[i][1]
				duplicates.append((f, open_rows[open_idx]))
			else:
				unique_fgp.append(f)
				
		for idx, o in enumerate(open_rows):
			if not idx in [i[1] for i in dup_indices]:
				unique_open.append(o)
		
		final_duplicates = []
		for d in duplicates:
			dup_dict = collections.OrderedDict()
			dup_dict['Source'] = 'FGP & Open Maps'
			dup_dict['Title'] = d[0]['Title']
			dup_dict['Description'] = d[0]['Description']
			if d[0]['Publisher'] == '':
				pub_str = d[1]['Publisher']
			else:
				pub_str = d[0]['Publisher']
			dup_dict['Publisher'] = pub_str
			dup_dict['FGP Topic Category'] = d[0]['FGP Topic Category']
			dup_dict['FGP Metadata URL'] = d[0]['FGP Metadata URL']
			dup_dict['FGP Metadata Standard'] = d[0]['FGP Metadata Standard']
			dup_dict['FGP Metadata Keywords'] = d[0]['FGP Metadata Keywords']
			dup_dict['Open Maps Topic Category'] = d[1]['Open Maps Topic Category']
			dup_dict['Open Maps Metadata URL'] = d[1]['Open Maps Metadata URL']
			dup_dict['Open Maps Metadata Standard'] = d[1]['Open Maps Metadata Standard']
			dup_dict['Open Maps Metadata Keywords'] = d[1]['Open Maps Metadata Keywords']
			final_duplicates.append(dup_dict)
		
		#answer = raw_input("Press enter...")
		
		return unique_fgp, unique_open, final_duplicates
	
	def get_format(self, format, name=''):
		valid_formats = ['JSON', 'CSV', 'XLS', 'XLSX', 'PDF', 'ZIP', 'SHP', 
						'TIFF', 'ASCII Grid', 'DOCX', 'WMS', 'KMZ', 'KML', 
						'REST', 'GDB']
	
		if format == 'other':
			for v in valid_formats:
				if name.find(v) > -1:
					return name
				
		return format
		
	def get_iso_info(self, mdata_soup):
	
		# Check if mdata_soup only contains 1 line
		line_check = str(mdata_soup).split('\n')
		
		#print "\nSoup length: %s" % len(line_check)
		if len(line_check) < 4: return None
	
		desc_str = bsoup.get_text(mdata_soup.find('gmd:abstract'))
					
		title_tag = mdata_soup.find('gmd:title')
		title_str = bsoup.get_text(title_tag.find('gco:characterstring'))
		
		# Get the publisher
		pub_str = ''
		roles = mdata_soup.find_all('gmd:role')
		for role in roles:
			role_str = bsoup.get_text(role)
			if role_str.find('publisher') > -1:
				# Get its previous sibling <gmd:organisationName>
				org_name = role.find_previous_sibling('gmd:organisationname')
				pub_str = bsoup.get_text(org_name)
				
		# Get the metadata standard
		mdata_tag = mdata_soup.find('gmd:metadatastandardname')
		mdata_stand = bsoup.get_text(mdata_tag.find('gco:characterstring'))
		
		# Get all md_keywords
		md_keys = mdata_soup.find_all('gmd:md_keywords')
		keywords = []
		thes_keywords = []
		for k in md_keys:
			# Determine if the keyword is a thesaurus word
			thes_tag = k.find('gmd:thesaurusname')
			# Get a list of keywords
			sub_keywords = k.find_all('gmd:keyword')
			for sub in sub_keywords:
				key_tag = sub.find('gco:characterstring')
				key_str = bsoup.get_text(key_tag)
				if thes_tag is not None:
					thes_keywords.append(key_str)
				else:
					keywords.append(key_str)
			
		# Get the topic category
		tc_gmd = mdata_soup.find('gmd:md_topiccategorycode')
		tc_str = bsoup.get_text(tc_gmd)
		# Capitalize the first letter only
		tc_str = tc_str[0].upper() + tc_str[1:]
		# Split topic category at capital letters
		tc_split = re.findall('[A-Z][^A-Z]*', tc_str)
		tc_words = [tc.lower() for tc in tc_split]
		tcwords_str = ' '.join(tc_words)
		tcwords_str = tcwords_str[0].upper() + tcwords_str[1:]
		#print "Topic Category: %s" % tc_str
		#answer = raw_input("Press enter...")
			
		out_dict = collections.OrderedDict()
			
		out_dict['Title'] = title_str
		out_dict['Description'] = desc_str
		out_dict['Publisher'] = pub_str
		out_dict['Standard'] = mdata_stand
		out_dict['Keywords'] = keywords
		out_dict['Topic Category'] = tcwords_str
		out_dict['Subject Thesaurus'] = thes_keywords
		
		return out_dict
		
	def get_fgp_data(self, query_url, mdatadir_url, word=None, org=None, 
						both=False):
	
		# Build the search URL and search the portal
		params = collections.OrderedDict()
		
		if word is not None:
			params['title_OR_abstract_OR_any'] = word.replace(" ", "%2B")
		if org is not None:
			params['orgNameCanada_eng'] = org.replace(" ", "+")

		srch_url = shared.get_post_query(query_url, params)
		
		print srch_url
		
		xml_soup = bsoup.get_xml_soup(srch_url)
		
		if not self.check_result(xml_soup, srch_url, "FGP Page"):
			return None
		
		results = xml_soup.find_all('metadata')
		
		print "\nNumber of results: %s" % len(results)
		
		out_lst = []
		
		for idx, res in enumerate(results):
		
			out_dict = collections.OrderedDict()
			
			msg = "Extracting %s of approximately %s results from '%s'" \
				 % (idx + 1, len(results), query_url)
			shared.print_oneliner(msg)
		
			ds_uuid = bsoup.get_text(res.find('uuid'))
			ds_id = bsoup.get_text(res.find('id'))
			#print "ds_uuid: %s" % ds_uuid
			
			mdata_xml_url = mdatadir_url + "?uuid=%s" % ds_uuid
				
			#print mdata_xml_url
			
			# Soup up the XML metadata
			mdata_xml_soup = bsoup.get_soup(mdata_xml_url)
			
			if self.check_result(mdata_xml_soup, output=False):				
				xml_fill = PatternFill(fill_type=None)
			
				mdata_info = self.get_iso_info(mdata_xml_soup)
				
				if mdata_info is None: continue
					
				#print keywords
				#print mdata_stand
				
				title_str = mdata_info['Title']
				desc_str = mdata_info['Description']
				pub_str = mdata_info['Publisher']
				mdata_stand = mdata_info['Standard']
				keywords = mdata_info['Keywords']
				tc_str = mdata_info['Topic Category']
				thes_words = mdata_info['Subject Thesaurus']
				
				#answer = raw_input("Press enter...")
				
				mdata_url = "https://gcgeo.gc.ca/geonetwork/metadata/eng/%s" % ds_uuid
			else:
				xml_fill = PatternFill("solid", fgColor="FF9966")
			
				mdata_url = "https://gcgeo.gc.ca/geonetwork/metadata/eng/%s" % ds_uuid
				
				#print mdata_url
				
				# Open the metadata page
				mdata_soup = bsoup.get_soup(mdata_url, True, delay=1)
				#mdata_soup = bsoup.get_soup(mdata_url) #, True, delay=1)
				
				# Get the title
				h1 = mdata_soup.find('h1', attrs={'id': 'wb-cont'})
				title_str = bsoup.get_text(h1)
				
				# Get the description
				pre_desc = mdata_soup.find('pre', attrs={'itemprop': 
											'description'})
				desc_str = bsoup.get_text(pre_desc)
				
				# Get the publisher
				contains = 'gmd:CI_ResponsibleParty/gmd:organisationName'
				span = bsoup.find_tag_by_id(mdata_soup, contains, 
												'span')
				tr = bsoup.get_parent(span, 'tr')
				tr_sib = tr.find_next_sibling('tr')
				pub_str = bsoup.get_text(tr_sib)
				
				# Get the metadata standard
				contains = 'gmd:MD_Metadata/gmd:metadataStandardName'
				span = bsoup.find_tag_by_id(mdata_soup, contains, 
												'span')
				tr = bsoup.get_parent(span, 'tr')
				tr_sib = tr.find_next_sibling('tr')
				mdata_stand = bsoup.get_text(tr_sib)
				
				keywords = []
				
				#print "pub_str: %s" % pub_str
				
				#answer = raw_input("Press enter..."
				
			if self.debug:
				print "Title: %s" % title_str
				print "Description: %s" % desc_str
				print "Publisher: %s" % pub_str
				print "Topic Category: %s" % tc_str
				print "Core Subject Thesaurus: %s" % ', '.join(thes_words)
				print "Metadata URL: %s" % mdata_url
				print "Metadata ID: %s" % ds_id
				print 'Metadata Standard: %s' % mdata_stand
				print 'Metadata Keywords: %s' % ', '.join(keywords)
				print 'Metadata XML URL: %s' % mdata_xml_url
				
			# Add all values to the CSV file object
			out_dict['Source'] = 'FGP'
			out_dict['Title'] = title_str
			out_dict['Description'] = desc_str
			out_dict['Publisher'] = pub_str
			if both:
				out_dict['FGP Topic Category'] = tc_str
				out_dict['FGP Core Subject Thesaurus'] = ', '.join(thes_words)
				out_dict['FGP Metadata URL'] = mdata_url
				out_dict['FGP Metadata ID'] = ds_id
				out_dict['FGP Metadata Standard'] = mdata_stand
				out_dict['FGP Metadata Keywords'] = ', '.join(keywords)
				out_dict['FGP Metadata XML URL'] = mdata_xml_url
			else:
				out_dict['Topic Category'] = tc_str
				out_dict['Core Subject Thesaurus'] = ', '.join(thes_words)
				out_dict['Metadata URL'] = mdata_url
				out_dict['Metadata ID'] = ds_id
				out_dict['Metadata Standard'] = mdata_stand
				out_dict['Metadata Keywords'] = ', '.join(keywords)
				out_dict['Metadata XML URL'] = mdata_xml_url

			out_lst.append(out_dict)
			
		return out_lst
		
		# # Get the soup for the query page results
		# res_soup = bsoup.get_soup(srch_url)
			
		# if not self.check_result(res_soup, srch_url, "FGP Page"):
			# return None

		# # Get the total number of records
		# strong = bsoup.find_tags_containing(res_soup, '1-', 'strong')
		# #self.url_out_f.write('FGP URL: %s\n' % srch_url)
		# if strong is None:
			# print "\nNo results found for '%s'." % word
			# return None
		
		# rec_total = int(bsoup.get_text(strong).split(' of ')[1])
		
		# print rec_total
		
		# record_count = 0
		
		# #print int(self.get_arg_val('start'))
		
		# out_lst = []
		
		# #start = int(self.get_arg_val('start'))
		# for pos in range(0, rec_total, 10):
			
			# # Get the start and end position for the URL
			# start_pos = pos + 1
			# end_pos = start_pos + 9
			
			# # Set the query URL
			# new_params = collections.OrderedDict()
			# if word is not None and not word == '':
				# new_params['title_OR_abstract_OR_any'] = word.replace(" ", "+")
			# if org is not None and not org == '':
				# new_params['orgNameCanada_eng'] = org.replace(" ", "+")
			# new_params['from'] = str(start_pos)
			# new_params['to'] = str(end_pos)
			
			# #print new_params
			
			# page_url = shared.get_post_query(query_url, new_params)
			
			# # Get the page soup
			# page_soup = bsoup.get_soup(page_url)
			
			# if not self.check_result(page_soup, page_url, 'FGP Portal'):
				# continue
			
			# # Get all the <a> with onclick commands
			# a_list = page_soup.find_all('a', attrs={'onclick': 'mdBackPageManager.setCurrentPagePosition()'})
			
			# print "\nNumber of results on page: %s" % len(a_list)
			
			# #print page_url
			
			# # Go through each <a>
			# for idx, a in enumerate(a_list):
			
				# out_dict = collections.OrderedDict()
			
				# record_count += 1
				# msg = "Extracting %s of approximately %s results from '%s'" \
					# % (record_count, rec_total, query_url)
				# shared.print_oneliner(msg)
				
				# # Get the href and use to get the dataset's ID
				# href = a['href']
				# parsed = href.split('/')
				# id = parsed[len(parsed) - 1]
				
				# mdata_xml_url = mdatadir_url + "?uuid=%s" % id
				
				# #print mdata_xml_url
				
				# # Soup up the XML metadata
				# mdata_xml_soup = bsoup.get_soup(mdata_xml_url)
				
				# if self.check_result(mdata_xml_soup, output=False):				
					# xml_fill = PatternFill(fill_type=None)
				
					# mdata_info = self.get_iso_info(mdata_xml_soup)
					
					# if mdata_info is None: continue
						
					# #print keywords
					# #print mdata_stand
					
					# title_str = mdata_info['Title']
					# desc_str = mdata_info['Description']
					# pub_str = mdata_info['Publisher']
					# mdata_stand = mdata_info['Standard']
					# keywords = mdata_info['Keywords']
					# tc_str = mdata_info['Topic Category']
					# thes_words = mdata_info['Subject Thesaurus']
					
					# #answer = raw_input("Press enter...")
					
					# mdata_url = "https://gcgeo.gc.ca/geonetwork/metadata/eng/%s" % id
				# else:
					# xml_fill = PatternFill("solid", fgColor="FF9966")
				
					# mdata_url = "https://gcgeo.gc.ca/geonetwork/metadata/eng/%s" % id
					
					# #print mdata_url
					
					# # Open the metadata page
					# mdata_soup = bsoup.get_soup(mdata_url, True, delay=1)
					# #mdata_soup = bsoup.get_soup(mdata_url) #, True, delay=1)
					
					# # Get the title
					# h1 = mdata_soup.find('h1', attrs={'id': 'wb-cont'})
					# title_str = bsoup.get_text(h1)
					
					# # Get the description
					# pre_desc = mdata_soup.find('pre', attrs={'itemprop': 
												# 'description'})
					# desc_str = bsoup.get_text(pre_desc)
					
					# # Get the publisher
					# contains = 'gmd:CI_ResponsibleParty/gmd:organisationName'
					# span = bsoup.find_tag_by_id(mdata_soup, contains, 
													# 'span')
					# tr = bsoup.get_parent(span, 'tr')
					# tr_sib = tr.find_next_sibling('tr')
					# pub_str = bsoup.get_text(tr_sib)
					
					# # Get the metadata standard
					# contains = 'gmd:MD_Metadata/gmd:metadataStandardName'
					# span = bsoup.find_tag_by_id(mdata_soup, contains, 
													# 'span')
					# tr = bsoup.get_parent(span, 'tr')
					# tr_sib = tr.find_next_sibling('tr')
					# mdata_stand = bsoup.get_text(tr_sib)
					
					# keywords = []
					
					# #print "pub_str: %s" % pub_str
					
					# #answer = raw_input("Press enter..."
					
				# # Add all values to the CSV file object
				# out_dict['Source'] = 'FGP'
				# out_dict['Title'] = title_str
				# out_dict['Description'] = desc_str
				# out_dict['Publisher'] = pub_str
				# if both:
					# out_dict['FGP Topic Category'] = tc_str
					# out_dict['FGP Core Subject Thesaurus'] = ', '.join(thes_words)
					# out_dict['FGP Metadata URL'] = mdata_url
					# out_dict['FGP Metadata ID'] = id
					# out_dict['FGP Metadata Standard'] = mdata_stand
					# out_dict['FGP Metadata Keywords'] = ', '.join(keywords)
					# out_dict['FGP Metadata XML URL'] = mdata_xml_url
				# else:
					# out_dict['Topic Category'] = tc_str
					# out_dict['Core Subject Thesaurus'] = ', '.join(thes_words)
					# out_dict['Metadata URL'] = mdata_url
					# out_dict['Metadata ID'] = id
					# out_dict['Metadata Standard'] = mdata_stand
					# out_dict['Metadata Keywords'] = ', '.join(keywords)
					# out_dict['Metadata XML URL'] = mdata_xml_url

				# out_lst.append(out_dict)
	
	def get_openmaps_data(self, main_url, word=None, ds_type=None, coll=None, 
							keyword=None, org=None, both=False):
		# Build the search URL and search the portal
		params = collections.OrderedDict()
		if word is not None and not word == '': params['q'] = word
		if ds_type is not None and not ds_type == '':
			params['portal_type'] = ds_type
		if coll is not None and not coll == '':
			params['collection'] = coll
		if keyword is not None and not keyword == '':
			params['keywords'] = keyword
		if org is not None and not org == '':
			params['organization'] = org

		query_url = shared.get_post_query(main_url, params)

		res_soup = bsoup.get_soup(query_url)
		
		#print query_url
		
		#self.url_out_f.write('Open Maps URL: %s\n' % query_url)
		
		if not self.check_result(res_soup, query_url, 'Canada Open Maps'):
			return None
		
		page_count = bsoup.get_page_count(res_soup, 'ul', ('class', 'pagination'), 'li')
		
		print "Page count: %s" % page_count
		
		record_count = 0
		special_chr = ""
		
		# Get the record total
		rec_strong = bsoup.find_tags_containing(res_soup, 'records found', 'strong')
		strong_text = bsoup.get_text(rec_strong)
		if strong_text == '':
			rec_strong = bsoup.find_tags_containing(res_soup, 'record found', 'strong')
			strong_text = bsoup.get_text(rec_strong)
			
		if strong_text == '':
			records_total = 0
		else:
			#print "Strong text: '%s'" % strong_text
			total_str = strong_text.split(' ')[0].replace(',', '')
			#print "Total string: %s" % total_str
			if total_str == 'No': return None
			records_total = int(total_str)
			
		out_lst = []
		
		for page in range(0, page_count):
			# Open each iteration of pages:
			if special_chr == "":
				# Some of the search entries can include "&" character
				#	but if an error is returned, use the "?" character instead
				try:
					page_url = "%s&page=%s" % (query_url, page + 1)
					special_chr = "&"
				except:
					page_url = "%s?page=%s" % (query_url, page + 1)
					special_chr = "?"
			else:
				page_url = "%s%spage=%s" % (query_url, special_chr, page + 1)
				
			#print "page_url: %s" % page_url

			# Create the soup object of the current page
			page_soup = bsoup.get_soup(page_url, silent=True)
				
			if not self.check_result(page_soup, page_url, 'Open Government Portal'): continue
				
			# Get a list of results
			results = page_soup.find_all('article')
			
			# Let the user know if there are no records
			if len(results) == 0 and record_count == 0:
				print "No records exist with the given search parameters."
				print "URL query sample: %s" % query_url
				return None
			
			for res in results:
				out_dict = collections.OrderedDict()
			
				record_count += 1
				msg = "Extracting %s of approximately %s results from '%s'" % (record_count, records_total, query_url)
				shared.print_oneliner(msg)
			
				# Get the title and its URL to get the ID
				title_h3 = res.find('h3', attrs={'class': 'panel-title'})
				#title_str = bsoup.get_text(title_h3)
				res_a = title_h3.a
				ds_url = res_a['href'].split('/')
				id = ds_url[len(ds_url) - 1]
				#print "res_a: %s" % res_a
				webpage_url = shared.get_anchor_url(res_a, page_url)
				
				mdata_url = 'https://open.canada.ca/data/en/dataset/%s' % id
				
				##############################################################
				# Getting the XML ISO metadata:
				mdata_iso_url = 'https://csw.open.canada.ca/geonetwork/srv/' \
								'csw?service=CSW&version=2.0.2&request=' \
								'GetRecordById&outputSchema=csw:IsoRecord&' \
								'ElementSetName=full&id=%s' % id
								
				#print "\n%s" % mdata_iso_url
								
				#mdata_iso = shared.get_json(mdata_json_url)
				mdata_xml_soup = bsoup.get_soup(mdata_iso_url)
				
				title_str = bsoup.get_text(res_a)
				if not self.check_result(mdata_xml_soup, mdata_iso_url, title_str): continue
								
				mdata_info = self.get_iso_info(mdata_xml_soup)
				
				if mdata_info is None:
					desc_str = ''
					pub_str = ''
					mdata_stand = ''
					keywords = ''
					tc_str = ''
				else:
					title_str = mdata_info['Title']
					desc_str = mdata_info['Description']
					pub_str = mdata_info['Publisher']
					mdata_stand = mdata_info['Standard']
					keywords = mdata_info['Keywords']
					tc_str = mdata_info['Topic Category']
				
				##############################################################
				# Getting the JSON metadata:
				
				# mdata_json_url = 'https://open.canada.ca/data/en/api/3/action/' \
									# 'package_show?id=%s' % id
				
				# mdata_json = shared.get_json(mdata_json_url)
				
				# title_str = bsoup.get_text(res_a)
				# if not self.check_result(mdata_json, mdata_json_url, title_str): continue
				
				# print mdata_json_url
				
				# answer = raw_input("Press enter...")
				
				# # print title_str
				
				# ds_info = mdata_json['result']
				
				# title_str = shared.clean_text(ds_info['title_translated']['en'])
				
				# # Get the final resource
				# resources = ds_info['resources']
				
				# desc_str = shared.clean_text(ds_info['notes_translated']['en'])
				
				# # Get the metadata standard
				
				
				# # Get the keywords
				# keywords = ds_info['keywords']['en']
				
				# formats = []
				# downloads = []
				# dates = []
				# for resrc in resources:
				
					# format = resrc['format']
				
					# if format.lower() == 'html' or \
						# format.lower() == 'xml':
						# continue
						
					# download_str = resrc['url']
					# downloads.append(download_str)
					
					# #if download_str.find('MapServer') > -1 and not format == 'wms':
					# #	format = 'ESRI REST'
					
					# format = self.get_format(format, resrc['name'])
					
					# formats.append(format)
				
					# #date_str = ''
					# #if 'date_published' in resrc:
					# #	date_str = resrc['date_published']
					# #	dates.append(date_str)
					
				# formats = list(set(formats))
					
				# #download_info = shared.get_download_text(formats, downloads, dates)
				# #download_str, access_str, date_str = download_info.split('|')
				
				# # Get the publisher
				# pub_str = ds_info['org_title_at_publication']['en']
				
				##############################################################
				
				# Add all values to the CSV file object
				out_dict['Source'] = 'Open Maps'
				out_dict['Title'] = title_str
				out_dict['Description'] = desc_str
				out_dict['Publisher'] = pub_str
				if both:
					out_dict['Open Maps Topic Category'] = tc_str
					out_dict['Open Maps Metadata URL'] = mdata_url
					out_dict['Open Maps Metadata Standard'] = mdata_stand
					out_dict['Open Maps Metadata Keywords'] = ', '.join(keywords)
				else:
					out_dict['Topic Category'] = tc_str
					out_dict['Metadata URL'] = mdata_url
					out_dict['Metadata Standard'] = mdata_stand
					out_dict['Metadata Keywords'] = ', '.join(keywords)

				out_lst.append(out_dict)
				
		return out_lst, query_url
		
	def extract_all(self):
	
		self.print_log("\nExtracting from %s" % self.pg_grp.get_title())
		
		self.print_title("Extracting All Portals")
		
		fgp_main_url = self.pg_grp.get_url('fgp_main_url')
		fgp_query_url = self.pg_grp.get_url('fgp_query_url')
		fgp_mdatadir_url = self.pg_grp.get_url('fgp_mdata_url')
		open_url = self.pg_grp.get_url('open_url')
		
		# Open file to output query URLs
		self.url_out_f = open('CA_Query_URLs.txt', 'a')
		
		word = self.get_arg_val('word')
		#ds_type = self.get_arg_val('ds_type')
		coll = self.get_arg_val('category')
		#keyword = self.get_arg_val('keyword')
		org_val = self.get_arg_val('organization')
		
		# Create a string to let the user know what parameters
		#	are being used
		out_str = "\nUsing the following parameters for extraction:"
		if word is not None and not word == '':
			out_str += '\nSearch Word: %s' % word
		if coll is not None and not coll == '':
			out_str += '\nCollection/Category: %s' % coll
		if org_val is not None and not org_val == '':
			out_str += '\nOrganization: %s\n' % org_val
		print out_str
			
		# Convert abbreviation to department for FGP
		fgp_org = org_val
		for idx, org_opt in enumerate(self.org_opts):
			url_tags = org_opt.get_urltags()
			if org_val.lower() in [o.lower() for o in url_tags]:
				if idx < 10:
					fgp_org = url_tags[0]
					break
				else:
					if word is None or word == '':
						print "\nConverting organization to search word."
						fgp_org = ''
						word = url_tags[0]
						break
						
		# Convert abbreviation to department for Open Maps
		open_org = org_val
		for idx, org_opt in enumerate(self.org_opts):
			url_tags = org_opt.get_urltags()
			if org_val.lower() in [o.lower() for o in url_tags]:
				open_org = url_tags[len(url_tags) - 1]
				break
				
		#header = [('Source', 50), ('Title', 100), ('Description', 100), 
		#			('Publisher', 60), ('FGP Topic Category', 30), 
		#			('FGP Metadata URL', 70), ('FGP Metadata Standard', 70), 
		#			('FGP Metadata Keywords', 50), ('Open Maps Topic Category', 30), 
		#			('Open Maps Metadata URL', 70), ('Open Maps Metadata Standard', 70), 
		#			('Open Maps Metadata Keywords', 50)]
		header = sh.get_header_info('ca_both')['xl']
					
		#('Topic Category', 30), ('Metadata ID', 40), 
		#			('Metadata Keywords', 50), ('Metadata URL', 70), 
		#			('Metadata XML URL', 70)

		pt_xl = sh.PT_XL(self, header=header, write_only=True, replace_ws=True, 
							ws_title=org_val.upper())
		
		# Get the FGP rows
		fgp_rows = self.get_fgp_data(fgp_query_url, fgp_mdatadir_url, word, 
									fgp_org, both=True)
		
		# Get the Open Maps rows
		open_info = self.get_openmaps_data(open_url, word, coll=coll, 
									org=open_org, both=True)
		if open_info is not None:
			open_rows, query_url = open_info	
		else:
			open_rows = []
									
		# Compare rows for duplicates
		unique_fgp, unique_open, duplicates = self.compare(fgp_rows, open_rows)
		
		for row in unique_fgp:
			for k, v in row.items():
				#print "%s: %s" % (k, v)
				pt_xl.add_item(k, v)
		
			pt_xl.write_row()
			
		pt_xl.write_row()
			
		for row in unique_open:
			for k, v in row.items():
				#print "%s: %s" % (k, v)
				pt_xl.add_item(k, v)
		
			pt_xl.write_row()
			
		#answer = raw_input("Press enter...")
			
		pt_xl.write_row()
			
		for row in duplicates:
			for k, v in row.items():
				#print "%s: %s" % (k, v)
				pt_xl.add_item(k, v)
		
			pt_xl.write_row()
					
		#err_ws_name = "Errors OGP%s" % sheet_word
		self.write_err_xl(pt_xl)
				
		pt_xl.save_file()
		
	def extract_fgp(self):
		'''
		**********************************************************************
		Extracts the datasets from the FGP portal.
		**********************************************************************
		Parameters in self.argmt:
			- word: The search word for the portal
			- start: The starting position of the datasets (debug only).
			- xl: Determines whether to export to Excel spreadsheet.
		'''
		
		self.print_log("\nExtracting from %s" % self.pg_grp.get_title())
		
		self.print_title("Extracting the Federal Geospatial Platform")
		
		main_url = self.pg_grp.get_url('main_url')
		query_url = self.pg_grp.get_url('query_url')
		mdatadir_url = self.pg_grp.get_url('mdata_url')
		
		#xl = self.get_arg_val('xl')
		#method = self.pg_grp.get_id()
		word = self.get_arg_val('word')
		org = self.get_arg('organization')
		test = self.get_arg_val('test')
		
		if test is not None:
			if test.lower().find('y') > -1:
				main_url = main_url.replace('gcgeo.gc.ca', 'test.gcgeo.gc.ca')
				query_url = query_url.replace('gcgeo.gc.ca', 'test.gcgeo.gc.ca')
				mdatadir_url = mdatadir_url.replace('gcgeo.gc.ca', 'test.gcgeo.gc.ca')
		
		# Get the organization value for FGP
		org_val = org.get_optname()
		
		# Create a string to let the user know what parameters
		#	are being used
		out_str = "\nUsing the following parameters for extraction:"
		if word is not None and not word == '':
			out_str += '\nSearch Word: %s' % word
		if org_val is not None and not org_val == '':
			out_str += '\nOrganization: %s\n' % org_val
		print out_str
		
		opts_lst = [opt for opt in self.get_arg('organization').get_opts()]
		
		# Convert abbreviation to department
		org = org_val
		if org_val is not None:
			for idx, org_opt in enumerate(opts_lst):
				url_tags = org_opt.get_urltags()
				if org_val.lower() in [o.lower() for o in url_tags]:
					if idx < 10:
						fgp_org = url_tags[0]
						break
					else:
						if word is None or word == '':
							print "\nConverting organization to search word."
							fgp_org = ''
							word = url_tags[0]
							break
		
		#header = [('Source', 50), ('Title', 100), ('Description', 100), 
		#			('Publisher', 60), ('Topic Category', 30), ('Metadata ID', 40), 
		#			('Metadata Keywords', 50), ('Metadata URL', 70), 
		#			('Metadata XML URL', 70)]
		
		if self.xl:
			header = sh.get_header_info('ca_fgp')['xl']
		else:
			header = sh.get_header_info('ca_fgp')['csv']
			
		if org_val is None or org_val == '':
			if word is None or word == '':
				sh_title = ''
			else:
				sh_title = word.replace(' ', '_')
		else:
			sh_title = self.get_arg_val('organization').upper()
		
		if self.xl:
			# pt_xl = sh.PT_XL(self, header=header, write_only=True, 
							# replace_ws=True, ws_title="FGP", 
							# fn_word=fn_out)		
			pt_xl = sh.PT_XL(self, header=header, write_only=True, 
							replace_ws=True, ws_title=sh_title)
		else:
			# Create the CSV file
			csv_fn = "FGP_%s_results" % fn_out
			#header = ['Title', 'Metadata URL', 'Publisher']
			pt_csv = sh.PT_CSV(csv_fn, self, header)
			pt_csv.open_csv()
		
		fgp_rows = self.get_fgp_data(query_url, mdatadir_url, word, org)
		if fgp_rows is not None:
			if self.xl:
				for row in fgp_rows:
					for k, v in row.items():
						pt_xl.add_item(k, v)
				
					pt_xl.write_row()
			else:
				for row in fgp_rows:
					for k, v in row.items():
						pt_csv.add(k, v)
				
					pt_csv.write_dataset()
					
		#err_ws_name = "Errors OGP%s" % sheet_word
		if self.xl: self.write_err_xl(pt_xl)
		#else: self.write_err_xl(pt_csv)
				
		if self.xl: pt_xl.save_file()
		else: pt_csv.close_csv()
	
	def extract_portal(self):
		'''
		**********************************************************************
		Extracts the datasets from the Canadian Open Government portal.
		**********************************************************************
		Parameters in self.argmt:
			- word: The search word for the portal.
			- ds_type: Filter the search by this portal type.
			- category: Filter the search by this collection.
			- keyword: Filter the search by this keyword.
			- xl: Determines whether to export to Excel spreadsheet.
		'''
		###########################################################################
		# Extract the Canadian Open Government Portal

		main_url = self.pg_grp.get_url('main_url')

		self.print_log("\nExtracting from %s" % self.pg_grp.get_title())
		
		self.print_title("Extracting Canada's Open Data Portal")
		
		#method = self.pg_grp.get_id()
		word = self.get_arg_val('word')
		ds_type = self.get_arg_val('ds_type')
		coll = self.get_arg_val('category')
		keyword = self.get_arg_val('keyword')
		org_val = self.get_arg_val('organization')
		
		# Create a string to let the user know what parameters
		#	are being used
		out_str = "\nUsing the following parameters for extraction:"
		if word is not None and not word == '':
			out_str += '\nSearch Word: %s' % word
		if ds_type is not None and not ds_type == '':
			out_str += '\nDataset Type: %s' % ds_type
		if coll is not None and not coll == '':
			out_str += '\nCollection/Category: %s' % coll
		if keyword is not None and not keyword == '':
			out_str += '\nKeyword: %s' % keyword
		if org_val is not None and not org_val == '':
			out_str += '\nOrganization: %s\n' % org_val
		print out_str
		
		# Convert abbreviation to department
		org = org_val
		org_opts = self.get_arg_opts('organization')
		for idx, org_opt in enumerate(org_opts):
			url_tags = org_opt.get_urltags()
			if org_val.lower() in [o.lower() for o in url_tags]:
				open_org = url_tags[len(url_tags) - 1]
				break
				
		#header = [('Source', 50), ('Title', 100), ('Description', 100), 
		#			('Publisher', 60), ('Metadata URL', 70)]
					
		header = sh.get_header_info('ca_portal')['xl']

		if self.xl:
			pt_xl = sh.PT_XL(self, header=header, write_only=True, replace_ws=True, 
							ws_title="Open Government Portal", 
							fn_word=org_val)
		else:
			# Create the CSV file
			csv_fn = "Portal_results"
			#header = ['Title', 'Metadata URL', 'Publisher']
			pt_csv = sh.PT_CSV(csv_fn, self)
			pt_csv.open_csv()

		open_rows, query_url = self.get_openmaps_data(main_url, word, ds_type, 
								coll, keyword, org)	
		if self.xl:
			for row in open_rows:
				for k, v in row.items():
					pt_xl.add_item(k, v)
			
				pt_xl.write_row()
		else:
			for row in open_rows:
				for k, v in row.items():
					pt_csv.add(k, v)
			
				pt_csv.write_dataset()
					
		#err_ws_name = "Errors OGP%s" % sheet_word
		if self.xl: self.write_err_xl(pt_xl)
		
		if self.xl:
			pt_xl.add_cell("Query URL: %s" % query_url)
			pt_xl.write_row()
		
		if self.xl: pt_xl.save_file()
		else: pt_csv.close_csv()

def main():
	# tool_list = ['REST', 'GIS']
	# tool_list = ['Ducks', 'GIS', 'FlySask2']
	try:
		ext = Extractor()
		pages = ext.get_pagelist()

		parser = argparse.ArgumentParser()
		parser.add_argument("-p", "--page", help="The page to extract: %s or all" % ', '.join(pages.keys()))
		# parser.add_argument("-w", "--word", help="The key word(s) to search for.")
		# parser.add_argument("-f", "--format", help="The format(s) to search for.")
		# parser.add_argument("-c", "--category", help="The category(ies) to search for.")
		# parser.add_argument("-d", "--downloadable", help="Determines wheter to get only downloadable datasets.")
		# parser.add_argument("-l", "--html", help="The HTML file to scrape (only for OpenData website).")
		parser.add_argument("-s", "--silent", action='store_true', help="If used, no extra parameters will be queried.")
		args = parser.parse_args()
		# print args.echo

		# print "province: " + str(args.province)
		# print "format: " + str(args.format)

		page = args.page
		# word = args.word
		# formats = args.format
		# html = args.html
		silent = args.silent
		# cats = args.category
		# downloadable = args.downloadable

		if page is None:
			answer = raw_input("Please enter the page you would like to use (%s or all): " % ', '.join(pages.keys()))
			if not answer == "":
				page = answer.lower()
			else:
				print "\nERROR: Please specify a web page."
				print "Exiting process."
				sys.exit(1)

		page = page.lower()

		print page

		ext.set_page(page)
		ext.run()

	except Exception, err:
		ext.print_log('ERROR: %s\n' % str(err))
		ext.print_log(traceback.format_exc())
		ext.close_log()


if __name__ == '__main__':
	sys.exit(main())
